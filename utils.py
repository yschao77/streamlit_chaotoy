import streamlit as st
import pandas as pd
import openpyxl
import hashlib
import datetime
import io
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# =========================================================================
# 🛠️ 🔴 全域攔截並修補 openpyxl 描述器核心驗證 Bug (Monkey Patch)
# =========================================================================
def apply_openpyxl_patch():
    try:
        import openpyxl.descriptors.base
        orig_set_attr = openpyxl.descriptors.base.Set.__set__

        def patched_set_attr(self, instance, value):
            if isinstance(value, str) and '-' in value:
                parts = value.split('-')
                value = parts[0] + ''.join(p.title() for p in parts[1:])
            
            try:
                orig_set_attr(self, instance, value)
            except ValueError:
                if hasattr(self, 'values'):
                    if isinstance(self.values, set):
                        self.values.add(value)
                    elif isinstance(self.values, tuple):
                        self.values = self.values + (value,)
                    elif isinstance(self.values, list):
                        self.values.append(value)
                orig_set_attr(self, instance, value)

        openpyxl.descriptors.base.Set.__set__ = patched_set_attr
    except Exception:
        pass

# =========================================================================
# 🛠️ check python-calamine
# =========================================================================
def check_calamine():
    try:
        import calamine
        return True
    except ImportError:
        return False

HAS_CALAMINE = check_calamine()

# =========================================================================
# 🌐 1. Google Drive 雲端連線初始化
# =========================================================================
@st.cache_resource
def init_drive_service():
    """讀取部署後設定在 Streamlit Secrets 的金鑰字典並建立雲端連線"""
    try:
        google_secrets = st.secrets["textkey"]
        credentials = service_account.Credentials.from_service_account_info(
            google_secrets,
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        return build('drive', 'v3', credentials=credentials)
    except Exception as e:
        st.error(f"❌ 無法從 Streamlit Secrets 中讀取 `textkey` 憑證。錯誤訊息: {str(e)}")
        st.info("💡 請確認您的 Secrets 設定格式是否正確。")
        st.stop()

# 建立全域 service 供下方工具使用
service = init_drive_service()

# =========================================================================
# 🔍 2. 雲端核心實戰工具與搜尋常式
# =========================================================================
@st.cache_data(ttl=3600)
def get_cached_gdrive_id(folder_id, file_name_keyword):
    try:
        query = f"'{folder_id}' in parents and name contains '{file_name_keyword}' and trashed = false"
        results = service.files().list(q=query, fields="files(id, name, modifiedTime)", pageSize=1).execute()
        files = results.get('files', [])
        if files:
            return files[0]['id'], files[0]['modifiedTime'], files[0]['name']
    except Exception:
        pass
    return None, None, None

def list_gdrive_files(folder_id):
    try:
        query = f"'{folder_id}' in parents and (mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or mimeType = 'application/vnd.ms-excel.sheet.macroEnabled.12') and trashed = false"
        results = service.files().list(q=query, fields="files(id, name, modifiedTime)").execute()
        files = results.get('files', [])
        files.sort(key=lambda x: x['name'], reverse=True)
        return files
    except Exception as e:
        st.error(f"掃描雲端資料夾失敗: {str(e)}")
        return []

def download_gdrive_file_to_bytes(file_id):
    request = service.files().get_media(fileId=file_id)
    file_stream = io.BytesIO()
    downloader = MediaIoBaseDownload(file_stream, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
    file_stream.seek(0)
    return file_stream

@st.cache_data(ttl=600, show_spinner="☁️ 正在從雲端載入檔案...")
def get_cached_gdrive_file_bytes(file_id):
    file_stream = download_gdrive_file_to_bytes(file_id)
    return file_stream.getvalue()

def upload_or_update_gdrive_file(folder_id, file_name, file_bytes, existing_file_id=None):
    file_name_str = str(file_name).lower()
    if file_name_str.endswith('.xlsm'):
        mime_type = 'application/vnd.ms-excel.sheet.macroEnabled.12'
    elif file_name_str.endswith('.xls'):
        mime_type = 'application/vnd.ms-excel'
    else:
        mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype=mime_type, resumable=True)
    
    if existing_file_id:
        service.files().update(fileId=existing_file_id, media_body=media, supportsAllDrives=True).execute()
        return existing_file_id
    else:
        st.error(f"❌ 拒絕建立新檔案【{file_name}】！為避免 Google 空間配額與權限錯誤，請先手動於雲端建立該檔案。")
        st.stop()

def format_gdrive_time(time_str):
    if not time_str: return "❌ 雲端檔案尚未建立/不存在"
    try:
        dt = datetime.datetime.strptime(time_str.split('.')[0], "%Y-%m-%dT%H:%M:%S")
        dt = dt + datetime.timedelta(hours=8)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except:
        return time_str

# =========================================================================
# 🧹 3. 資料清洗與輔助工具
# =========================================================================
def calculate_md5(file_bytes):
    return hashlib.md5(file_bytes).hexdigest()

def clean_barcode(val):
    if pd.isna(val): return ""
    s = str(val).strip().replace(" ", "").replace("\xa0", "")
    if s.lower() == "nan" or s == "": return ""
    if s.endswith(".0"): s = s[:-2]
    if "e+" in s.lower() or "e" in s.lower():
        try: s = f"{float(s):.0f}"
        except: pass
    return s

def process_smart_headers(df_raw, header_row_idx):
    if header_row_idx >= 5 and len(df_raw) > 5:
        row_5 = df_raw.iloc[4].fillna("").astype(str).str.strip()
        row_6 = df_raw.iloc[5].fillna("").astype(str).str.strip()
        new_cols = []
        for idx in range(len(df_raw.columns)):
            c5 = row_5.iloc[idx] if idx < len(row_5) else ""
            c6 = row_6.iloc[idx] if idx < len(row_6) else ""
            if any(k in c5 for k in ["訂購", "CTN", "內盒", "數量", "單價"]) and c5 != c6 and c6 != "":
                new_cols.append(f"{c5}{c6}")
            elif c6 != "": new_cols.append(c6)
            else: new_cols.append(c5)
        return new_cols
    else:
        return [str(col).strip() for col in df_raw.iloc[header_row_idx].fillna("")]

# =========================================================================
# 📦 4. 資料庫共用載入常式
# =========================================================================
@st.cache_data(ttl=600)
def load_shopee_data(file_id):
    """延遲載入並快取蝦皮主表資料"""
    if not file_id: return pd.DataFrame(columns=["檔案名稱", "md5", "匯入時間"]), pd.DataFrame()
    try:
        shopee_bytes = download_gdrive_file_to_bytes(file_id)
        with pd.ExcelFile(shopee_bytes) as shopee_xls:
            df_hist = pd.read_excel(shopee_xls, "匯入檔案") if "匯入檔案" in shopee_xls.sheet_names else pd.DataFrame(columns=["檔案名稱", "md5", "匯入時間"])
            df_list = pd.read_excel(shopee_xls, "蝦皮商品列表") if "蝦皮商品列表" in shopee_xls.sheet_names else pd.DataFrame()
        return df_hist, df_list
    except Exception:
        return pd.DataFrame(columns=["檔案名稱", "md5", "匯入時間"]), pd.DataFrame()


# =========================================================================
# 💾 5. 核心資料庫讀寫與勾稽常式
# =========================================================================
import openpyxl

def save_to_master_xlsm(sheets_dict, file_id, folder_id, file_name):
    if not file_id:
        st.error(f"❌ 雲端找不到核心總表檔案")
        return False
    try:
        master_bytes = download_gdrive_file_to_bytes(file_id)
        wb = openpyxl.load_workbook(master_bytes, keep_vba=True)
        for sheet_name, df in sheets_dict.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.delete_rows(1, ws.max_row + 1)
            else:
                ws = wb.create_sheet(sheet_name)
            ws.append(list(df.columns))
            
            barcode_col_idx = list(df.columns).index("條碼") + 1 if "條碼" in df.columns else None
            for row_idx, row in enumerate(df.itertuples(index=False), start=2):
                cleaned_row = []
                for col_idx, x in enumerate(row):
                    if pd.isna(x): cleaned_row.append("")
                    elif barcode_col_idx and (col_idx + 1) == barcode_col_idx: cleaned_row.append(str(x).strip().split('.')[0])
                    else: cleaned_row.append(x)
                ws.append(cleaned_row)
                if barcode_col_idx: ws.cell(row=row_idx, column=barcode_col_idx).number_format = '@'
                    
        out_buf = io.BytesIO()
        wb.save(out_buf)
        upload_or_update_gdrive_file(folder_id, file_name or "麗嬰採購產品總表.xlsm", out_buf.getvalue(), existing_file_id=file_id)
        return True
    except Exception as e:
        st.error(f"❌ 寫入雲端資料庫發生錯誤: {str(e)}")
        return False

def save_to_shopee_master_xlsm(sheets_dict, file_id, folder_id, file_name):
    try:
        if file_id:
            shopee_bytes = download_gdrive_file_to_bytes(file_id)
            wb = openpyxl.load_workbook(shopee_bytes, keep_vba=True)
        else:
            wb = openpyxl.Workbook()
            
        for sheet_name, df in sheets_dict.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.delete_rows(1, ws.max_row + 1)
            else:
                ws = wb.create_sheet(sheet_name)
            ws.append(list(df.columns))
            for row in df.itertuples(index=False):
                cleaned_row = [" " if pd.isna(x) else x for x in row]
                ws.append(cleaned_row)
                
        out_buf = io.BytesIO()
        wb.save(out_buf)
        upload_or_update_gdrive_file(
            folder_id, 
            file_name or "蝦皮賣場商品列表.xlsm", 
            out_buf.getvalue(), 
            existing_file_id=file_id
        )
        return True
    except Exception as e:
        st.error(f"❌ 寫入雲端蝦皮資料庫發生錯誤: {str(e)}")
        return False

def run_cross_matching(df):
    if df.empty: return df
    df['備註'] = df['備註'].astype(str).apply(lambda x: "" if x == "nan" or x == "None" else x)
    records = df.to_dict('records')
    from collections import defaultdict
    barcode_groups = defaultdict(list)
    for idx, row in enumerate(records):
        b_str = str(row.get('條碼', '')).strip().split('.')[0]
        if b_str and b_str not in ["", "0", "nan", "None"]:
            barcode_groups[b_str].append(idx)
            
    for b_str, indices in barcode_groups.items():
        if len(indices) > 1:
            for check_idx in indices:
                check_row = records[check_idx]
                name_str = str(check_row.get('名稱', '')).strip()
                existing_remark = str(check_row.get('備註', '')).strip()
                if existing_remark and not any(k in existing_remark for k in ["條碼重複", "名稱不同", "零售價不同"]):
                    continue
                    
                try: price_val = float(check_row.get('零售價', 0)) if pd.notna(check_row.get('零售價', 0)) else 0
                except: price_val = 0
                
                for comp_idx in indices:
                    if check_idx == comp_idx: continue
                    comp_row = records[comp_idx]
                    comp_name_str = str(comp_row.get('名稱', '')).strip()
                    try: comp_price_val = float(comp_row.get('零售價', 0)) if pd.notna(comp_row.get('零售價', 0)) else 0
                    except: comp_price_val = 0
                    uid_str = str(comp_row.get('UID', '未知')).strip()
                    
                    if name_str != comp_name_str:
                        records[check_idx]['備註'] = f"與 UID: {uid_str} 條碼重複, 名稱不同"
                        break
                    elif price_val != comp_price_val:
                        records[check_idx]['備註'] = f"與 UID: {uid_str} 條碼重複, 名稱相同, 零售價不同"
                        break
    return pd.DataFrame(records)

@st.cache_data(ttl=600)
def load_master_data(file_id):
    if not file_id: return None, None, None, None, None, 3473
    try:
        master_bytes = download_gdrive_file_to_bytes(file_id)
        with pd.ExcelFile(master_bytes) as xls:
            df_total = pd.read_excel(xls, "麗嬰國際產品總表")
            df_history = pd.read_excel(xls, "已處理採購單")
            df_delete_log = pd.read_excel(xls, "刪除紀錄") if "刪除紀錄" in xls.sheet_names else pd.DataFrame(columns=["UID", "名稱", "條碼", "零售價", "備註", "匯入檔名", "刪除時間"])
            df_meta = pd.read_excel(xls, "metadata")
            all_sheets = xls.sheet_names
            
        if "條碼" in df_total.columns:
            df_total['條碼'] = df_total['條碼'].astype(str).str.strip().str.split('.').str[0]
        if "條碼" in df_delete_log.columns:
            df_delete_log['條碼'] = df_delete_log['條碼'].astype(str).str.strip().str.split('.').str[0]
            
        if not df_history.empty:
            df_history.columns = [str(col).strip().lower() for col in df_history.columns]
            df_history = df_history.loc[:, ~df_history.columns.duplicated()].copy()
            standard_cols = ["檔案名稱", "md5", "匯入時間"]
            if len(df_history.columns) >= 3:
                df_history.columns = standard_cols + list(df_history.columns[3:])
            else:
                df_history = pd.DataFrame(columns=standard_cols)
        else:
            df_history = pd.DataFrame(columns=["檔案名稱", "md5", "匯入時間"])
            
        current_max_uid = int(df_meta.iloc[0, 0]) if not df_meta.empty else 3473
        return df_total, df_history, df_delete_log, df_meta, all_sheets, current_max_uid
    except Exception as e:
        st.error(f"🔴 讀取雲端主資料庫失敗。錯誤: {str(e)}")
        return None, None, None, None, None, 3473