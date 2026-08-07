import streamlit as st
import pandas as pd
import openpyxl
import hashlib
import datetime
import io
import os
import json
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# 設定網頁標題與寬度
st.set_page_config(page_title="麗嬰與蝦皮商務數據情報中心", page_icon="📊", layout="wide")

# =========================================================================
# 🛠️ 🔴 全域攔截並修補 openpyxl 描述器核心驗證 Bug (Monkey Patch)
# =========================================================================
try:
    import openpyxl.descriptors.base
    orig_set_attr = openpyxl.descriptors.base.Set.__set__

    def patched_set_attr(self, instance, value):
        if isinstance(value, str) and '-' in value:
            parts = value.split('-')
            value = parts[0] + ''.join(p.title() for p in parts[1:])
        
        try:
            orig_set_attr(self, instance, value)
        except ValueError:
            if hasattr(self, 'values'):
                if isinstance(self.values, set):
                    self.values.add(value)
                elif isinstance(self.values, tuple):
                    self.values = self.values + (value,)
                elif isinstance(self.values, list):
                    self.values.append(value)
            orig_set_attr(self, instance, value)

    openpyxl.descriptors.base.Set.__set__ = patched_set_attr
except Exception:
    pass
    
# =========================================================================
# 🛠️ check python-calamine
# =========================================================================
try:
    import calamine
    HAS_CALAMINE = True
except ImportError:
    HAS_CALAMINE = False

# =========================================================================
# 🌐 1. Google Drive 雲端資料夾 ID 定義與權限初始化
# =========================================================================
ID_PROD_FOLDER = "1NtMAYb-SvdH6XMmqB5G07ttB-NuWCqDV"          # 商品列表 資料夾 ID
ID_PRICE_SUMMARY_FOLDER = "1ZM4MscX0UO6rUHjKv-mN5fKDwxg53maZ" # 價格統整表 資料夾 ID
ID_SHOPEE_FOLDER = "17eiGnXyU4KwNS6IR5bubBPti46SKXMH0"        # 蝦皮商品清單 資料夾 ID
ID_HISTORY_INWARD_FOLDER = "1ZQ7x4BdRc6BJlURxQ61JqDKrKF7h_vSH"# 歷史入庫單 資料夾 ID
ID_BASE_FOLDER = "1HjMt8z8DXlqGhSqe50_hDR3f4LpVLK_w"          # 麗嬰採購統整 資料夾 ID

@st.cache_resource
def init_drive_service():
    """讀取部署後設定在 Streamlit Secrets 的金鑰字典並建立雲端連線"""
    try:
        google_secrets = st.secrets["textkey"]
        credentials = service_account.Credentials.from_service_account_info(
            google_secrets,
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        return build('drive', 'v3', credentials=credentials)
    except Exception as e:
        st.error(f"❌ 無法從 Streamlit Secrets 中讀取 `textkey` 憑證。錯誤訊息: {str(e)}")
        st.info("💡 請確認您的 Secrets 設定格式是否正確。")
        st.stop()

service = init_drive_service()

# =========================================================================
# 🔍 2. 雲端核心實戰工具與搜尋常式
# =========================================================================
@st.cache_data(ttl=3600)  # 優化：快取雲端檔案搜尋結果 1 小時
def get_cached_gdrive_id(folder_id, file_name_keyword):
    """在指定的 Google Drive 資料夾中，根據名稱搜尋檔案並返回其 ID、修改時間、完整檔名"""
    try:
        query = f"'{folder_id}' in parents and name contains '{file_name_keyword}' and trashed = false"
        results = service.files().list(q=query, fields="files(id, name, modifiedTime)", pageSize=1).execute()
        files = results.get('files', [])
        if files:
            return files[0]['id'], files[0]['modifiedTime'], files[0]['name']
    except Exception:
        pass
    return None, None, None

def list_gdrive_files(folder_id):
    """列出雲端資料夾內的所有 Excel 檔案"""
    try:
        query = f"'{folder_id}' in parents and (mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or mimeType = 'application/vnd.ms-excel.sheet.macroEnabled.12') and trashed = false"
        results = service.files().list(q=query, fields="files(id, name, modifiedTime)").execute()
        files = results.get('files', [])
        files.sort(key=lambda x: x['name'], reverse=True)
        return files
    except Exception as e:
        st.error(f"掃描雲端資料夾失敗: {str(e)}")
        return []

def download_gdrive_file_to_bytes(file_id):
    """將雲端檔案下載至記憶體中"""
    request = service.files().get_media(fileId=file_id)
    file_stream = io.BytesIO()
    downloader = MediaIoBaseDownload(file_stream, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
    file_stream.seek(0)
    return file_stream

@st.cache_data(ttl=600, show_spinner="☁️ 正在從雲端載入檔案...")
def get_cached_gdrive_file_bytes(file_id):
    """快取雲端檔案的純 Bytes 內容，確保 Streamlit 快取安全且避免重複下載"""
    file_stream = download_gdrive_file_to_bytes(file_id)
    return file_stream.getvalue()  # 💡 加上 .getvalue()，將串流轉化為純 bytes！

def upload_or_update_gdrive_file(folder_id, file_name, file_bytes, existing_file_id=None):
    """【強制覆寫優化版】動態判斷 mimetype，避免 Google Drive 鎖定檔案"""
    
    # 💡 1. 根據附檔名動態判定正確的 MIME Type
    file_name_str = str(file_name).lower()
    if file_name_str.endswith('.xlsm'):
        mime_type = 'application/vnd.ms-excel.sheet.macroEnabled.12'
    elif file_name_str.endswith('.xls'):
        mime_type = 'application/vnd.ms-excel'
    else:
        mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
    # 2. 封裝上傳內容
    media = MediaIoBaseUpload(
        io.BytesIO(file_bytes), 
        mimetype=mime_type, 
        resumable=True
    )
    
    if existing_file_id:
        # 🟢 只准執行 Update，並維持原檔案擁有者權限
        service.files().update(
            fileId=existing_file_id, 
            media_body=media, 
            supportsAllDrives=True
        ).execute()
        return existing_file_id
    else:
        # 🔴 防呆攔截：阻斷 Create 行為
        st.error(f"❌ 拒絕建立新檔案【{file_name}】！為避免 Google 空間配額與權限錯誤，請先手動於雲端建立該檔案。")
        st.stop()

def format_gdrive_time(time_str):
    if not time_str:
        return "❌ 雲端檔案尚未建立/不存在"
    try:
        dt = datetime.datetime.strptime(time_str.split('.')[0], "%Y-%m-%dT%H:%M:%S")
        dt = dt + datetime.timedelta(hours=8)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except:
        return time_str

def calculate_md5(file_bytes):
    return hashlib.md5(file_bytes).hexdigest()

# -------------------------------------------------------------------------
# 動態偵測各核心主表的雲端 ID ＆ 最後修改時間 (使用 Cache 加速)
# -------------------------------------------------------------------------
ID_MASTER_FILE, TIME_MASTER, NAME_MASTER = get_cached_gdrive_id(ID_BASE_FOLDER, "麗嬰採購產品總表")
ID_LOCAL_PROD, TIME_PROD, NAME_PROD = get_cached_gdrive_id(ID_PROD_FOLDER, "商品列表")
ID_SHOPEE_MASTER, TIME_SHOPEE, NAME_SHOPEE = get_cached_gdrive_id(ID_SHOPEE_FOLDER, "蝦皮賣場商品列表")
ID_PRICE_SUMMARY, TIME_SUMMARY, NAME_SUMMARY = get_cached_gdrive_id(ID_PRICE_SUMMARY_FOLDER, "商品蝦皮麗嬰價格統整表")

# =========================================================================
# ⚙️ 3. 核心資料庫讀寫雲端常式
# =========================================================================
def save_to_master_xlsm(sheets_dict):
    if not ID_MASTER_FILE:
        st.error(f"❌ 雲端找不到核心總表檔案 【麗嬰採購產品總表.xlsm】")
        return False
    try:
        master_bytes = download_gdrive_file_to_bytes(ID_MASTER_FILE)
        wb = openpyxl.load_workbook(master_bytes, keep_vba=True)
        for sheet_name, df in sheets_dict.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.delete_rows(1, ws.max_row + 1)
            else:
                ws = wb.create_sheet(sheet_name)
            ws.append(list(df.columns))
            
            barcode_col_idx = list(df.columns).index("條碼") + 1 if "條碼" in df.columns else None
            for row_idx, row in enumerate(df.itertuples(index=False), start=2):
                cleaned_row = []
                for col_idx, x in enumerate(row):
                    if pd.isna(x): cleaned_row.append("")
                    elif barcode_col_idx and (col_idx + 1) == barcode_col_idx: cleaned_row.append(str(x).strip().split('.')[0])
                    else: cleaned_row.append(x)
                ws.append(cleaned_row)
                if barcode_col_idx: ws.cell(row=row_idx, column=barcode_col_idx).number_format = '@'
                    
        out_buf = io.BytesIO()
        wb.save(out_buf)
        upload_or_update_gdrive_file(ID_BASE_FOLDER, NAME_MASTER or "麗嬰採購產品總表.xlsm", out_buf.getvalue(), existing_file_id=ID_MASTER_FILE)
        return True
    except Exception as e:
        st.error(f"❌ 寫入雲端資料庫發生錯誤: {str(e)}")
        return False

def save_to_shopee_master_xlsm(sheets_dict):
    global ID_SHOPEE_MASTER
    try:
        if ID_SHOPEE_MASTER:
            shopee_bytes = download_gdrive_file_to_bytes(ID_SHOPEE_MASTER)
            wb = openpyxl.load_workbook(shopee_bytes, keep_vba=True)
        else:
            wb = openpyxl.Workbook()
            
        for sheet_name, df in sheets_dict.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.delete_rows(1, ws.max_row + 1)
            else:
                ws = wb.create_sheet(sheet_name)
            ws.append(list(df.columns))
            for row in df.itertuples(index=False):
                cleaned_row = [" " if pd.isna(x) else x for x in row]
                ws.append(cleaned_row)
                
        out_buf = io.BytesIO()
        wb.save(out_buf)
        ID_SHOPEE_MASTER = upload_or_update_gdrive_file(
            ID_SHOPEE_FOLDER, 
            NAME_SHOPEE or "蝦皮賣場商品列表.xlsm", 
            out_buf.getvalue(), 
            existing_file_id=ID_SHOPEE_MASTER
        )
        return True
    except Exception as e:
        st.error(f"❌ 寫入雲端蝦皮資料庫發生錯誤: {str(e)}")
        return False

def run_cross_matching(df):
    if df.empty: return df
    df['備註'] = df['備註'].astype(str).apply(lambda x: "" if x == "nan" or x == "None" else x)
    records = df.to_dict('records')
    from collections import defaultdict
    barcode_groups = defaultdict(list)
    for idx, row in enumerate(records):
        b_str = str(row.get('條碼', '')).strip().split('.')[0]
        if b_str and b_str not in ["", "0", "nan", "None"]:
            barcode_groups[b_str].append(idx)
            
    for b_str, indices in barcode_groups.items():
        if len(indices) > 1:
            for check_idx in indices:
                check_row = records[check_idx]
                name_str = str(check_row.get('名稱', '')).strip()
                existing_remark = str(check_row.get('備註', '')).strip()
                if existing_remark and not any(k in existing_remark for k in ["條碼重複", "名稱不同", "零售價不同"]):
                    continue
                    
                try: price_val = float(check_row.get('零售價', 0)) if pd.notna(check_row.get('零售價', 0)) else 0
                except: price_val = 0
                
                for comp_idx in indices:
                    if check_idx == comp_idx: continue
                    comp_row = records[comp_idx]
                    comp_name_str = str(comp_row.get('名稱', '')).strip()
                    try: comp_price_val = float(comp_row.get('零售價', 0)) if pd.notna(comp_row.get('零售價', 0)) else 0
                    except: comp_price_val = 0
                    uid_str = str(comp_row.get('UID', '未知')).strip()
                    
                    if name_str != comp_name_str:
                        records[check_idx]['備註'] = f"與 UID: {uid_str} 條碼重複, 名稱不同"
                        break
                    elif price_val != comp_price_val:
                        records[check_idx]['備註'] = f"與 UID: {uid_str} 條碼重複, 名稱相同, 零售價不同"
                        break
    return pd.DataFrame(records)

# =========================================================================
# ⚙️ 4. 全域雲端資料庫載入與初始化安全驗證 (優化：全面快取與延遲載入)
# =========================================================================
@st.cache_data(ttl=600)
def load_master_data(file_id):
    """延遲載入並快取麗嬰主表資料"""
    if not file_id: return None, None, None, None, None, 3473
    try:
        master_bytes = download_gdrive_file_to_bytes(file_id)
        # 優化：此處若不寫入，僅供查詢，建議預設使用 openpyxl 讀取保留相容性，但查詢頁面可改用 calamine
        with pd.ExcelFile(master_bytes) as xls:
            df_total = pd.read_excel(xls, "麗嬰國際產品總表")
            df_history = pd.read_excel(xls, "已處理採購單")
            df_delete_log = pd.read_excel(xls, "刪除紀錄") if "刪除紀錄" in xls.sheet_names else pd.DataFrame(columns=["UID", "名稱", "條碼", "零售價", "備註", "匯入檔名", "刪除時間"])
            df_meta = pd.read_excel(xls, "metadata")
            all_sheets = xls.sheet_names
            
        if "條碼" in df_total.columns:
            df_total['條碼'] = df_total['條碼'].astype(str).str.strip().str.split('.').str[0]
        if "條碼" in df_delete_log.columns:
            df_delete_log['條碼'] = df_delete_log['條碼'].astype(str).str.strip().str.split('.').str[0]
            
        if not df_history.empty:
            df_history.columns = [str(col).strip().lower() for col in df_history.columns]
            df_history = df_history.loc[:, ~df_history.columns.duplicated()].copy()
            standard_cols = ["檔案名稱", "md5", "匯入時間"]
            if len(df_history.columns) >= 3:
                df_history.columns = standard_cols + list(df_history.columns[3:])
            else:
                df_history = pd.DataFrame(columns=standard_cols)
        else:
            df_history = pd.DataFrame(columns=["檔案名稱", "md5", "匯入時間"])
            
        current_max_uid = int(df_meta.iloc[0, 0]) if not df_meta.empty else 3473
        return df_total, df_history, df_delete_log, df_meta, all_sheets, current_max_uid
    except Exception as e:
        st.error(f"🔴 讀取雲端主資料庫失敗。錯誤: {str(e)}")
        return None, None, None, None, None, 3473

@st.cache_data(ttl=600)
def load_shopee_data(file_id):
    """延遲載入並快取蝦皮主表資料"""
    if not file_id: return pd.DataFrame(columns=["檔案名稱", "md5", "匯入時間"]), pd.DataFrame()
    try:
        shopee_bytes = download_gdrive_file_to_bytes(file_id)
        with pd.ExcelFile(shopee_bytes) as shopee_xls:
            df_hist = pd.read_excel(shopee_xls, "匯入檔案") if "匯入檔案" in shopee_xls.sheet_names else pd.DataFrame(columns=["檔案名稱", "md5", "匯入時間"])
            df_list = pd.read_excel(shopee_xls, "蝦皮商品列表") if "蝦皮商品列表" in shopee_xls.sheet_names else pd.DataFrame()
        return df_hist, df_list
    except Exception:
        return pd.DataFrame(columns=["檔案名稱", "md5", "匯入時間"]), pd.DataFrame()

if 'inward_input_df' not in st.session_state:
    st.session_state['inward_input_df'] = pd.DataFrame([{"國際條碼": "", "數量": 1}])

# --- 條碼格式終極清洗函數 ---
def clean_barcode(val):
    """強制清洗條碼：去除空格、網頁空白字元、還原 E+ 科學記號、去除浮點數後綴"""
    if pd.isna(val): 
        return ""
    s = str(val).strip().replace(" ", "").replace("\xa0", "")
    if s.lower() == "nan" or s == "": 
        return ""
    if s.endswith(".0"): 
        s = s[:-2]
    # 還原 Pandas 誤讀的科學記號 (如 4.71E+12)
    if "e+" in s.lower() or "e" in s.lower():
        try:
            s = f"{float(s):.0f}"
        except:
            pass
    return s

# --- 智慧複合表頭處理函數 ---
def process_smart_headers(df_raw, header_row_idx):
    """完美重現 VBA 邏輯：當第5列有特定字眼時，與第6列合併為新標頭"""
    if header_row_idx >= 5 and len(df_raw) > 5:
        row_5 = df_raw.iloc[4].fillna("").astype(str).str.strip()
        row_6 = df_raw.iloc[5].fillna("").astype(str).str.strip()
        new_cols = []
        for idx in range(len(df_raw.columns)):
            c5 = row_5.iloc[idx] if idx < len(row_5) else ""
            c6 = row_6.iloc[idx] if idx < len(row_6) else ""
            
            # 若第5列包含採購關鍵字，且與第6列不同，則動態結合 (例如: 內盒 + 12 = 內盒12)
            if any(k in c5 for k in ["訂購", "CTN", "內盒", "數量", "單價"]) and c5 != c6 and c6 != "":
                new_cols.append(f"{c5}{c6}")
            elif c6 != "":
                new_cols.append(c6)
            else:
                new_cols.append(c5)
        return new_cols
    else:
        return [str(col).strip() for col in df_raw.iloc[header_row_idx].fillna("")]

# ==========================================
# 🧭 5. 側邊欄：導覽控制台
# ==========================================
st.sidebar.markdown("## 🏢 進銷存中央管理系統")
st.sidebar.write("---")

main_module = st.sidebar.selectbox(
    "🎯 請選擇核心管理模組：",
    ["🏪 sitegiant 電商整合管理", "📦 商品蝦皮麗嬰統整管理"]
)
st.sidebar.write("") 

if "商品蝦皮麗嬰統整管理" in main_module:
    st.sidebar.markdown("### 🛠️ 整合合併轉換功能")
    sub_page = st.sidebar.radio(
        "請選擇執行項目：",
        ["🧠 PowerQuery 執行三表整合", "⚖️ 麗嬰商品表合併和與審核", "📈 蝦皮商品清單轉換", "📊 PowerQuery 三表整合歷史紀錄", "🔍 商品清單紀錄查詢", "🔍 麗嬰商品總表數據查詢"],
        index=3
    )
else:
    st.sidebar.markdown("### 🌐 sitegiant 電商整合管理")
    sub_page = st.sidebar.radio(
        "請選擇執行項目：",
        ["🔀 sitegiant 採購入庫單格式轉換", "📜 sitegiant 歷史入庫單紀錄", "📦 SiteGiant 批量新增UPC", "📋 採購單待處理"],
        index=0
    )

# ==========================================
# 🖥️ 6. 各子分頁功能邏輯處理
# ==========================================
st.title(f"{sub_page}")
st.info(f"目前導覽路徑： {main_module} ➔ {sub_page}")
st.write("---")

# =========================================================================
# 🖥️ 7. 全域獨立功能: 三表 PowerQuery 整合、計算財務指標
# =========================================================================
def run_powerquery_and_update_gdrive(df_to_save=None):
    """
    全域獨立功能：執行三表 PowerQuery 整合、計算財務指標，並強制 Update 覆寫至雲端統整表。
    若傳入 df_to_save，則直接回寫雲端，確保 WYSIWYG (所見即所得)。
    """
    if not (ID_LOCAL_PROD and ID_SHOPEE_MASTER and ID_MASTER_FILE):
        st.error("❌ 缺少核心資料主檔案 ID，無法啟動三表整合！")
        return False
        
    try:
        # 若沒有傳入預覽的 df，則啟動從頭計算
        if df_to_save is None:
            # 1. 跨表讀取數據
            engine_kw = {"engine": "calamine"} if HAS_CALAMINE else {}
            df_liying = pd.read_excel(download_gdrive_file_to_bytes(ID_MASTER_FILE), sheet_name="麗嬰國際產品總表", **engine_kw)
            df_p = pd.read_excel(download_gdrive_file_to_bytes(ID_LOCAL_PROD), sheet_name=0, **engine_kw)
            df_s = pd.read_excel(download_gdrive_file_to_bytes(ID_SHOPEE_MASTER), sheet_name="蝦皮商品列表", **engine_kw)
            
            # 2. 資料清洗與標準化
            df_liying['條碼'] = df_liying['條碼'].astype(str).str.strip().str.split('.').str[0]
            df_p["自定義編碼"] = df_p["自定義編碼"].astype(str).str.strip().str.split('.').str[0]
            df_s["iSKU"] = df_s["iSKU"].astype(str).str.strip().str.split('.').str[0]
            
            # 🌟 修復 1：若 df_p 已有「商品名稱」，先更名為「內部商品名稱」，避免 Merge 產生 _x, _y 後綴
            if "商品名稱" in df_p.columns:
                df_p = df_p.rename(columns={"商品名稱": "內部商品名稱"})
            
            # 3. 模擬 PowerQuery 進行多表 Merge 關聯
            df_merge1 = pd.merge(df_p, df_s[["商品名稱","iSKU", "GTIN", "價格"]], left_on="自定義編碼", right_on="iSKU", how="left")
            df_merge1 = df_merge1.rename(columns={"商品名稱": "蝦皮商品名稱", "GTIN": "蝦皮GTIN", "價格": "蝦皮售價"})
            df_merge1["c"] = df_merge1["c"].astype(str).str.strip().str.split('.').str[0]
            
            df_final = pd.merge(df_merge1, df_liying[["條碼", "零售價", "含稅"]], left_on="c", right_on="條碼", how="left")
            df_final = df_final.rename(columns={"零售價": "麗嬰零售價", "含稅": "麗嬰批發含稅價", "條碼": "麗嬰條碼"})
            df_final["麗嬰商品"] = df_final["麗嬰條碼"].apply(lambda x: None if pd.isna(x) else "v")
            
            # 4. 財務與稅款指標動態計算
            for c in ["蝦皮售價", "麗嬰零售價", "麗嬰批發含稅價"]:
                df_final[c] = pd.to_numeric(df_final[c], errors='coerce')
                
            df_final["麗嬰零售八折"] = df_final["麗嬰零售價"] * 0.8
            df_final["麗嬰八折比蝦皮貴"] = df_final.apply(lambda r: "v" if (pd.notna(r["麗嬰零售八折"]) and pd.notna(r["蝦皮售價"]) and r["麗嬰零售八折"] > r["蝦皮售價"]) else None, axis=1)
            df_final["麗嬰未稅價"] = df_final["麗嬰批發含稅價"].apply(lambda x: round(x / 1.05, 2) if pd.notna(x) else None)
            df_final["麗嬰稅款"] = df_final.apply(lambda r: round(r["麗嬰批發含稅價"] - r["麗嬰未稅價"], 2) if (pd.notna(r["麗嬰批發含稅價"]) and pd.notna(r["麗嬰未稅價"])) else None, axis=1)
            
            # 清除不必要的欄位並寫入全域狀態
            df_to_save = df_final.drop(columns=["iSKU"], errors="ignore")
            st.session_state['pq_result'] = df_to_save
            
        # 5. 將結果轉為記憶體二進位流並準備覆寫雲端 (100% 確保寫入的是我們整理好的 df_to_save)
        output_stream = io.BytesIO()
        with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
            df_to_save.to_excel(writer, index=False, sheet_name="商品蝦皮麗嬰價格統整表")
        output_stream.seek(0)
        
        # 尋找雲端現有的「商品蝦皮麗嬰價格統整表」檔案 ID
        existing_summary_id, _, _ = get_cached_gdrive_id(ID_PRICE_SUMMARY_FOLDER, "商品蝦皮麗嬰價格統整表")
        
        if existing_summary_id:
            upload_or_update_gdrive_file(
                folder_id=ID_PRICE_SUMMARY_FOLDER, 
                file_name="商品蝦皮麗嬰價格統整表.xlsx", 
                file_bytes=output_stream.getvalue(), 
                existing_file_id=existing_summary_id
            )
            
            if "gdrive_id_cache" in st.session_state:
                cache_key = f"{ID_PRICE_SUMMARY_FOLDER}_商品蝦皮麗嬰價格統整表"
                if cache_key in st.session_state["gdrive_id_cache"]:
                    del st.session_state["gdrive_id_cache"][cache_key]
            return True
        else:
            st.error("❌ 雲端不存在『商品蝦皮麗嬰價格統整表』空白主表檔案，無法執行覆寫更新。")
            return False
            
    except Exception as e:
        st.error(f"❌ 自動化整合或回寫雲端時發生異常: {str(e)}")
        return False

# -------------------------------------------------------------------------
# 子功能 1：📊 三表整合歷史
# -------------------------------------------------------------------------
if sub_page == "📊 PowerQuery 三表整合歷史紀錄":
    st.subheader("🔄 三表整合歷史紀錄追蹤")
    hist_pq_files = list_gdrive_files(ID_PRICE_SUMMARY_FOLDER)
    if not hist_pq_files:
        st.warning(f"💡 提示：目前雲端資料夾內尚無任何歷史檔案，請至『🧠 PowerQuery 三表整合』執行新建轉換。")
    else:
        file_options = {f['name']: f['id'] for f in hist_pq_files}
        selected_pq_file = st.selectbox("🎯 請選擇欲調閱的歷史整合報告：", list(file_options.keys()))
        
        if selected_pq_file:
            try:
                target_id = file_options[selected_pq_file]
                file_bytes = download_gdrive_file_to_bytes(target_id)
                # 優化：純讀取歷史紀錄，嘗試使用 calamine 加速
                df_pq_view = pd.read_excel(file_bytes, engine="calamine" if HAS_CALAMINE else None)
                st.markdown(f"📄 **目前調閱雲端檔案**：`{selected_pq_file}` ｜ 📊 **資料總項數**：`{len(df_pq_view)} 筆`")
                st.dataframe(df_pq_view, use_container_width=True)
                st.download_button(label="🔄 重新下載此歷史整合表 (.xlsx)", data=file_bytes.getvalue(), file_name=selected_pq_file, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"❌ 讀取雲端備份檔案失敗: {str(e)}")

# -------------------------------------------------------------------------
# 子功能 2：🧠 PowerQuery 三表整合 (加強下載報表與顯示雲端統整表最後修改時間)
# -------------------------------------------------------------------------
elif sub_page == "🧠 PowerQuery 執行三表整合":
    st.subheader("🔍 三表數據追蹤")
    
    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("📦 商品列表 (商品iSKU清單)", "已對接" if ID_LOCAL_PROD else "❌ 未偵測到")
        st.caption(f"📅 最後修改時間: \n`{format_gdrive_time(TIME_PROD)}`")
    with c2:
        st.metric("🧡 蝦皮資料庫主表", "已對接" if ID_SHOPEE_MASTER else "❌ 未偵測到")
        st.caption(f"📅 最後修改時間: \n`{format_gdrive_time(TIME_SHOPEE)}`")
    with c3:
        st.metric("🧸 麗嬰產品總表", "已對接" if ID_MASTER_FILE else "❌ 未偵測到")
        st.caption(f"📅 最後修改時間: \n`{format_gdrive_time(TIME_MASTER)}`")

    st.write("---")

    # 🌟 全新加入：在執行前先到雲端抓取「商品蝦皮麗嬰價格統整表」的目前狀態與最後修改時間
    st.subheader("📊 雲端『商品蝦皮麗嬰價格統整表』當前狀態")
    existing_summary_id, existing_summary_time, _ = get_cached_gdrive_id(ID_PRICE_SUMMARY_FOLDER, "商品蝦皮麗嬰價格統整表")
    
    if existing_summary_id:
        st.info(f"🟢 雲端已存在統整表檔案 ｜ 📅 最後修改時間：`{format_gdrive_time(existing_summary_time)}`")
    else:
        st.warning("⚠️ 雲端目前尚未建立『商品蝦皮麗嬰價格統整表』，回寫時系統將會自動全新建立。")

    st.write("---")

    if st.button("🛠️ 啟動三表整合與財務指標計算", type="primary", use_container_width=True):
        if not (ID_LOCAL_PROD and ID_SHOPEE_MASTER and ID_MASTER_FILE):
            st.error("❌ 無法啟動三表整合！請確認雲端對應資料夾內是否缺少必要的核心資料主檔案。")
        else:
            with st.spinner("正在由雲端載入數據流並進行大數據跨表計算..."):
                try:
                    engine_kw = {"engine": "calamine"} if HAS_CALAMINE else {}
                    df_liying = pd.read_excel(download_gdrive_file_to_bytes(ID_MASTER_FILE), sheet_name="麗嬰國際產品總表", **engine_kw)
                    df_p = pd.read_excel(download_gdrive_file_to_bytes(ID_LOCAL_PROD), sheet_name=0, **engine_kw)
                    df_s = pd.read_excel(download_gdrive_file_to_bytes(ID_SHOPEE_MASTER), sheet_name="蝦皮商品列表", **engine_kw)
                    
                    df_liying['條碼'] = df_liying['條碼'].astype(str).str.strip().str.split('.').str[0]
                    df_p["自定義編碼"] = df_p["自定義編碼"].astype(str).str.strip().str.split('.').str[0]
                    df_s["iSKU"] = df_s["iSKU"].astype(str).str.strip().str.split('.').str[0]
                    
                    # 🌟 修復 1：若 df_p 已有「商品名稱」，先更名避免 Merge 產生 _x, _y 後綴
                    if "商品名稱" in df_p.columns:
                        df_p = df_p.rename(columns={"商品名稱": "內部商品名稱"})
                    
                    df_merge1 = pd.merge(df_p, df_s[["商品名稱","iSKU", "GTIN", "價格"]], left_on="自定義編碼", right_on="iSKU", how="left")
                    df_merge1 = df_merge1.rename(columns={"商品名稱": "蝦皮商品名稱", "GTIN": "蝦皮GTIN", "價格": "蝦皮售價"})
                    df_merge1["c"] = df_merge1["c"].astype(str).str.strip().str.split('.').str[0]
                    
                    df_final = pd.merge(df_merge1, df_liying[["條碼", "零售價", "含稅"]], left_on="c", right_on="條碼", how="left")
                    df_final = df_final.rename(columns={"零售價": "麗嬰零售價", "含稅": "麗嬰批發含稅價", "條碼": "麗嬰條碼"})
                    df_final["麗嬰商品"] = df_final["麗嬰條碼"].apply(lambda x: None if pd.isna(x) else "v")
                    
                    for c in ["蝦皮售價", "麗嬰零售價", "麗嬰批發含稅價"]:
                        df_final[c] = pd.to_numeric(df_final[c], errors='coerce')
                        
                    df_final["麗嬰零售八折"] = df_final["麗嬰零售價"] * 0.8
                    df_final["麗嬰八折比蝦皮貴"] = df_final.apply(lambda r: "v" if (pd.notna(r["麗嬰零售八折"]) and pd.notna(r["蝦皮售價"]) and r["麗嬰零售八折"] > r["蝦皮售價"]) else None, axis=1)
                    df_final["麗嬰未稅價"] = df_final["麗嬰批發含稅價"].apply(lambda x: round(x / 1.05, 2) if pd.notna(x) else None)
                    df_final["麗嬰稅款"] = df_final.apply(lambda r: round(r["麗嬰批發含稅價"] - r["麗嬰未稅價"], 2) if (pd.notna(r["麗嬰批發含稅價"]) and pd.notna(r["麗嬰未稅價"])) else None, axis=1)
                    
                    st.session_state['pq_result'] = df_final.drop(columns=["iSKU"], errors="ignore")
                    st.success("🎉 三表 PowerQuery 交叉聯結與財務指標計算整合完成！")
                except Exception as e:
                    st.error(f"❌ 錯誤: {str(e)}")

    # ── 當有整合結果存在時，顯示報表預覽、本地下載功能、以及雲端回寫機制 ──
    if 'pq_result' in st.session_state and st.session_state['pq_result'] is not None:
        df_result = st.session_state['pq_result']
        
        st.subheader("📋 整合聯結情報報表輸出預覽")
        st.markdown(f"📊 **目前整合結果資料總項數**：`{len(df_result)} 筆`")
        st.dataframe(df_result, use_container_width=True)
        
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            towrite_pq = io.BytesIO()
            with pd.ExcelWriter(towrite_pq, engine='openpyxl') as writer:
                df_result.to_excel(writer, index=False, sheet_name="PowerQuery三表整合")
            st.download_button(
                label="📥 匯出並下載此三表整合交叉比對表 (.xlsx)", 
                data=towrite_pq.getvalue(), 
                file_name=f"三表整合比對結果_{datetime.date.today().strftime('%Y%m%d')}.xlsx", 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
        with col_btn2:
            if st.button("🔄 執行：將整合結果回寫並更新至雲端『商品蝦皮麗嬰價格統整表』", type="secondary", use_container_width=True):
                with st.spinner("💾 正在覆寫更新雲端現有統整表檔案..."):
                    # ── 🌟 修復 2：直接把畫面的 df_result 丟給它存檔，杜絕重新計算產生的差異 ──
                    if run_powerquery_and_update_gdrive(df_to_save=df_result):
                        st.success("✅ 雲端統整表已成功同步覆寫更新！")
                        st.info("💡 重新整理頁面後，上方將會顯示最新的修改時間。")
                        
# -------------------------------------------------------------------------
# 子功能 3：🔍 麗嬰商品總表數據查詢 (支援條碼、庫存SKU多筆查詢與 Calamine 加速)
# -------------------------------------------------------------------------
elif sub_page == "🔍 麗嬰商品總表數據查詢":
    st.subheader("📋 麗嬰採購產品總表資料庫分頁動態檢視")
    
    _, _, _, _, all_sheets, _ = load_master_data(ID_MASTER_FILE)
    
    if all_sheets:
        view_sheets = [s for s in all_sheets if s != "麗嬰產品新採購單"]
        selected_sheet = st.selectbox("請選擇數據分頁：", view_sheets)
        
        search_mode = st.radio("🎯 請選擇查詢模式：", ["多筆條碼價格查詢", "多筆庫存SKU查詢", "模糊關鍵字搜尋"], horizontal=True)
        
        try:
            # 💡 使用快取與 calamine 高速讀取
            file_bytes = get_cached_gdrive_file_bytes(ID_MASTER_FILE)
            engine_kw = {"engine": "calamine"} if HAS_CALAMINE else {}
            df_view = pd.read_excel(io.BytesIO(file_bytes), sheet_name=selected_sheet, **engine_kw)
            
            # 模式 1：多筆條碼價格查詢
            if search_mode == "多筆條碼價格查詢":
                if "條碼" not in df_view.columns:
                    st.warning(f"⚠️ 當前選擇的分頁 【{selected_sheet}】 內部不含「條碼」欄位。")
                else:
                    df_view['條碼'] = df_view['條碼'].apply(clean_barcode)
                    barcode_paste = st.text_area("📋 請貼上多筆國際條碼 (換行、空格或逗號隔開)：", height=120, placeholder="例如：\n4711234567890\n4711234567891")
                    
                    if barcode_paste.strip():
                        import re
                        cleaned_barcodes = [clean_barcode(t) for t in re.split(r'[\n,\s]+', barcode_paste) if t.strip()]
                        cleaned_barcodes = [b for b in cleaned_barcodes if b]
                        
                        if cleaned_barcodes:
                            df_result = df_view[df_view['條碼'].isin(cleaned_barcodes)].copy()
                            st.success(f"🔍 查詢完畢！成功比對出 {len(df_result)} 筆商品資料。")
                            
                            important_cols = ["條碼", "名稱", "零售價", "含稅"]
                            display_cols = [c for c in important_cols if c in df_result.columns] + [c for c in df_result.columns if c not in important_cols]
                            st.dataframe(df_result[display_cols], use_container_width=True)
                            
                            towrite_query = io.BytesIO()
                            with pd.ExcelWriter(towrite_query, engine='openpyxl') as writer:
                                df_result[display_cols].to_excel(writer, index=False, sheet_name="條碼查詢結果")
                            st.download_button("📥 下載本次條碼查詢結果報表 (.xlsx)", data=towrite_query.getvalue(), file_name=f"條碼批次查詢結果_{datetime.date.today().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            # 模式 2：多筆庫存SKU查詢 (新增功能)
            elif search_mode == "多筆庫存SKU查詢":
                # 自動尋找可能代表庫存SKU的欄位名稱 (如 自定義編碼 或 庫存SKU)
                sku_col_candidate = next((c for c in ["庫存SKU", "自定義編碼", "商品編號", "貨號"] if c in df_view.columns), None)
                
                if not sku_col_candidate:
                    st.warning(f"⚠️ 當前選擇的分頁 【{selected_sheet}】 找不到對應的 SKU 欄位。")
                else:
                    df_view[sku_col_candidate] = df_view[sku_col_candidate].astype(str).str.strip().str.split('.').str[0]
                    sku_paste = st.text_area(f"📋 請貼上多筆【{sku_col_candidate}】（換行、空格或逗號隔開）：", height=120, placeholder="例如：\nSKU001\nSKU002")
                    
                    if sku_paste.strip():
                        import re
                        cleaned_skus = [t.strip() for t in re.split(r'[\n,\s]+', sku_paste) if t.strip()]
                        if cleaned_skus:
                            df_result = df_view[df_view[sku_col_candidate].isin(cleaned_skus)].copy()
                            st.success(f"🔍 查詢完畢！成功比對出 {len(df_result)} 筆商品資料。")
                            st.dataframe(df_result, use_container_width=True)
                            
                            towrite_query = io.BytesIO()
                            with pd.ExcelWriter(towrite_query, engine='openpyxl') as writer:
                                df_result.to_excel(writer, index=False, sheet_name="SKU查詢結果")
                            st.download_button("📥 下載本次SKU查詢結果報表 (.xlsx)", data=towrite_query.getvalue(), file_name=f"SKU批次查詢結果_{datetime.date.today().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            # 模式 3：模糊關鍵字搜尋
            elif search_mode == "模糊關鍵字搜尋":
                search_term = st.text_input("🔍 快速搜尋關鍵字 (支援條碼、品名、貨號模糊比對)：", placeholder="輸入搜尋內容...")
                st.metric(label=f"📊 【{selected_sheet}】當前總資料筆數", value=f"{len(df_view)} 筆")
                
                if search_term:
                    search_mask = df_view.astype(str).apply(lambda x: x.str.contains(search_term, case=False, na=False)).any(axis=1)
                    st.dataframe(df_view[search_mask], use_container_width=True)
                else:
                    st.dataframe(df_view, use_container_width=True)
                    
        except Exception as e:
            st.error(f"❌ 讀取分頁數據失敗: {str(e)}")
# -------------------------------------------------------------------------
# 子功能 4：⚖️ 麗嬰商品表合併和與審核 
# -------------------------------------------------------------------------
elif sub_page == "⚖️ 麗嬰商品表合併和與審核":
    st.subheader("🧸 麗嬰採購單一鍵導入與審核系統")
    
    # 讀取主檔 
    try:
        # 1. 檢查檔案 ID 是否存在
        if not ID_MASTER_FILE:
            st.error("❌ 錯誤：未設定雲端主資料庫 ID (ID_MASTER_FILE)！請檢查 Streamlit Secrets 或設定檔。")
            st.stop()

        # 2. 呼叫讀取函式
        master_data_result = load_master_data(ID_MASTER_FILE)

        # 3. 驗證回傳結果是否正常 (防呆判斷)
        if master_data_result is None or not isinstance(master_data_result, (tuple, list)):
            st.error("❌ 讀取雲端主資料庫失敗：無法從 Google Drive 載入資料，請確認 Service Account 權限與檔案 ID 是否正確。")
            st.stop()

        # 4. 解構賦值
        df_total, df_history, df_delete_log, df_meta, _, current_max_uid = master_data_result

        # 5. 確保核心 Dataframe 載入成功
        if df_total is None:
            st.error("❌ 讀取失敗：主資料庫內容 (df_total) 為空或無法正常解析，請檢查 Excel 工作表名稱與格式！")
            st.stop()

    except Exception as e:
        st.error(f"❌ 讀取雲端主資料庫時發生未預期的例外錯誤：{str(e)}")
        st.stop()

    # 若歷史表單中尚未建立 "狀態" 欄位，先進行初始化防錯
    if "狀態" not in df_history.columns:
        df_history["狀態"] = ""

    if 'merge_success_msg' in st.session_state:
        st.success(st.session_state['merge_success_msg'])
        st.markdown("### ⚡ 歸檔後後續自動化推薦操作")
        st.info("💡 採購單已成功存入麗嬰總表！直接點擊下方按鈕， PowerQuery 整合並自動更新雲端統整表。")
        if st.button("🚀 三表資料整合並自動回寫更新至雲端『商品蝦皮麗嬰價格統整表』", type="primary", use_container_width=True):
            with st.spinner("⏳ 正在跨資料庫調閱核心數據、執行大數據 VLOOKUP 計算並回寫雲端..."):
                if run_powerquery_and_update_gdrive():
                    st.success("🎯 狂賀！三表整合『商品蝦皮麗嬰價格統整表』已在雲端同步覆寫更新完畢！")
                    del st.session_state['merge_success_msg']   
        st.write("---")

    uploaded_files = st.file_uploader("📥 選擇採購單 Excel (可多選批次上傳)", type=["xlsx", "xls", "xlsm"], accept_multiple_files=True, key="main_merge_files")
    
    if uploaded_files:
        if st.button("🚀 開始一鍵合併到麗嬰總表", type="primary"):
            success_count = dup_count = no_barcode_count = anomaly_count = 0
            new_rows, history_records = [], []
            valid_dfs_to_merge = []
            
            # 將總表轉換為記憶體高速字典，並事先清洗總表條碼作為 Key
            master_dict = {}
            for idx, row in df_total.iterrows():
                b_key = clean_barcode(row.get('條碼', ''))
                if b_key and b_key != "0":
                    master_dict[b_key] = row
            
            # 確保歷史檔案 MD5 紀錄格式正確
            history_md5_list = df_history['md5'].astype(str).tolist() if 'md5' in df_history.columns else []

            with st.spinner("⏳ 正在執行：智慧表頭解析、防錯清洗與狀態紀錄..."):
                for file in uploaded_files:
                    file_bytes = file.read()
                    file_md5 = calculate_md5(file_bytes)
                    filename = file.name
                    now_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    # 1. 狀態判定：[重複檔案]
                    if file_md5 in history_md5_list:
                        dup_count += 1
                        history_records.append({"檔案名稱": filename, "md5": file_md5, "匯入時間": now_time, "狀態": "[重複檔案]"})
                        st.warning(f"⚠️ 檔案 `{filename}` 為重複檔案，已登記狀態並略過。")
                        continue
                        
                    try:
                        # 讀取前15列尋找真正的標頭列
                        df_src_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, nrows=15)
                        header_row = 0
                        for idx, row in df_src_raw.iterrows():
                            if row.astype(str).str.contains("條碼|國際條碼|EAN|Barcode|BARCODE").any():
                                header_row = idx
                                break
                        
                        # 智慧複合表頭處理
                        smart_columns = process_smart_headers(df_src_raw, header_row)
                        
                        # 完整讀取數據層，套用智慧標頭
                        df_src = pd.read_excel(io.BytesIO(file_bytes), skiprows=header_row + 1, header=None)
                        if df_src.empty or len(smart_columns) != len(df_src.columns):
                            df_src = pd.read_excel(io.BytesIO(file_bytes), header=header_row)
                        else:
                            df_src.columns = smart_columns

                        # 進行欄位標準化對齊
                        rename_dict = {}
                        for col in df_src.columns:
                            col_clean = str(col).strip()
                            if col_clean in ["名稱", "品名", "商品名稱", "中文", "中文品名", "品名規格", "Description"]: rename_dict[col] = "名稱"
                            elif col_clean in ["條碼", "國際條碼", "EAN", "Barcode", "BARCODE", "條碼型號", "JAN CODE"]: rename_dict[col] = "條碼"
                            elif col_clean in ["零售價", "建議售價", "單價", "定價", "售價", "Price"]: rename_dict[col] = "零售價"
                            elif col_clean in ["料品編號","商品編號", "貨號", "產品編號", "Item No", "ITEM"]: rename_dict[col] = "商品編號"
                            elif col_clean in ["內盒", "Inner", "INNER"]: rename_dict[col] = "內盒"
                            elif col_clean in ["CTN", "外箱", "箱入數", "Carton", "外箱數"]: rename_dict[col] = "CTN"
                            elif col_clean in ["CTN訂購含稅價", "CTN含稅價", "CTN含稅", "外箱含稅價", "外箱進價"]: rename_dict[col] = "CTN含稅"
                            elif col_clean in ["內盒訂購含稅價", "內盒含稅價", "內盒含稅", "內盒進價", "含稅", "含稅價", "進價"]: rename_dict[col] = "含稅"
                        df_src = df_src.rename(columns=rename_dict)
                        
                        # 2. 狀態判定：[無條碼欄位] (找不到條碼標題)
                        if "條碼" not in df_src.columns:
                            no_barcode_count += 1
                            history_records.append({"檔案名稱": filename, "md5": file_md5, "匯入時間": now_time, "狀態": "[無條碼欄位]"})
                            st.error(f"❌ 檔案 `{filename}` 找不到條碼欄位，已登記狀態並略過。")
                            continue
                            
                        # 條碼格式終極清洗
                        df_src['條碼'] = df_src['條碼'].apply(clean_barcode)
                        
                        # 3. 狀態判定：[無條碼欄位] (清洗後內容全空)
                        df_valid = df_src[df_src['條碼'] != ""]
                        if df_valid.empty:
                            no_barcode_count += 1
                            history_records.append({"檔案名稱": filename, "md5": file_md5, "匯入時間": now_time, "狀態": "[無條碼欄位]"})
                            st.error(f"❌ 檔案 `{filename}` 內無有效條碼數據，已登記狀態並略過。")
                            continue
                            
                        # 檔案驗證通過，準備進入合併
                        valid_dfs_to_merge.append({
                            "file_name": filename,
                            "file_md5": file_md5,
                            "df": df_valid
                        })
                        
                    except Exception as e:
                        # 解析失敗，歸類為無條碼/結構錯亂
                        st.error(f"❌ 檔案 【{filename}】 解析失敗: {str(e)}")
                        no_barcode_count += 1
                        history_records.append({"檔案名稱": filename, "md5": file_md5, "匯入時間": now_time, "狀態": "[無條碼欄位]"})
                        
            # 若有驗證通過的檔案，執行核心勾稽邏輯
            if valid_dfs_to_merge:
                with st.spinner("⏳ 正在執行雲端 UID 流水號繼承與雙向交叉審核比對..."):
                    for item in valid_dfs_to_merge:
                        df_item = item["df"]
                        filename = item["file_name"]
                        file_md5 = item["file_md5"]
                        now_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                        for _, src_row in df_item.dropna(subset=['條碼']).iterrows():
                            barcode = str(src_row['條碼'])
                            if barcode in ["", "0"]: continue
                            
                            price = round(float(src_row.get('零售價', 0)), 2) if pd.notna(src_row.get('零售價', 0)) else 0
                            name = str(src_row.get('名稱', '')).strip()
                            
                            match_found = False
                            need_insert = False
                            remark_note = ""
                            display_barcode = barcode

                            # 異常條碼重複勾稽與「雙向審核標記」
                            if barcode in master_dict:
                                match_found = True
                                master_row = master_dict[barcode]
                                target_uid = master_row.get('UID', '未知')
                                m_name = str(master_row.get('名稱', '')).strip()
                                
                                try: 
                                    m_price = float(master_row.get('零售價', 0)) if pd.notna(master_row.get('零售價', 0)) else 0
                                except: 
                                    m_price = 0
                                
                                if name and m_name != name: 
                                    need_insert = True
                                    anomaly_count += 1
                                    remark_note = f"與 UID: {target_uid} 條碼重複, 名稱不同"
                                    display_barcode = f"🔴 {barcode}"
                                elif price > 0 and m_price != price: 
                                    need_insert = True
                                    anomaly_count += 1
                                    remark_note = f"與 UID: {target_uid} 條碼重複, 名稱相同, 零售價不同"
                                    display_barcode = f"🟢 {barcode}"
                            
                            if (not match_found) or (match_found and need_insert):
                                current_max_uid += 1
                                new_uid = f"UID-{int(current_max_uid):06d}"
                                
                                row_data = {col: "" for col in df_total.columns if col != 'move'}
                                for col in df_item.columns:
                                    if col in row_data:
                                        val = src_row[col]
                                        row_data[col] = "" if pd.isna(val) else val
                                
                                row_data["UID"] = new_uid
                                row_data["條碼"] = display_barcode
                                row_data["名稱"] = name
                                if "零售價" in row_data: row_data["零售價"] = price
                                row_data["備註"] = remark_note
                                row_data["匯入檔名"] = filename
                                
                                new_rows.append(row_data)
                                master_dict[barcode] = row_data

                        # 4. 狀態判定：[已匯入] (成功處理完畢)
                        success_count += 1
                        history_records.append({"檔案名稱": filename, "md5": file_md5, "匯入時間": now_time, "狀態": "[已匯入]"})
                        history_md5_list.append(file_md5)
                    
            # 整合所有歷史狀態紀錄與新資料回寫雲端
            if new_rows: 
                df_total = pd.concat([df_total, pd.DataFrame(new_rows)], ignore_index=True)
            if history_records: 
                df_history = pd.concat([df_history, pd.DataFrame(history_records)], ignore_index=True)
            
            df_meta.iloc[0, 0] = current_max_uid
            df_total = run_cross_matching(df_total)
            
            # ⭐ 注意這裡將回寫的 Sheet 名稱改為 "已處理採購單"
            if save_to_master_xlsm({"麗嬰國際產品總表": df_total, "已處理採購單": df_history, "metadata": df_meta}):
                load_master_data.clear()
                
                report_msg = f"🎉 成功完成狀態登記與資料同步！\n\n✅ [已匯入]: {success_count} 份\n🔁 [重複檔案]: {dup_count} 份\n⚠️ [無條碼欄位]: {no_barcode_count} 份"
                if anomaly_count > 0:
                    report_msg += f"\n\n🚨 注意：本次匯入發現 **{anomaly_count}** 筆異常衝突商品，已自動為您加上 🔴 🟢 標記於備註欄！"
                
                st.session_state['merge_success_msg'] = report_msg
                st.rerun()
    

    st.write("---")
    st.subheader("⚠️ 條碼重複與衝突即時審核控制台")
    # --- 下方審核控制台邏輯維持不變 ---
    if not df_total.empty:
        if 'move' in df_total.columns: df_total = df_total.drop(columns=['move'])
        df_total.insert(0, 'move', False)
        df_total['move'] = df_total['move'].astype(bool)
        
        # 使用清洗後的條碼進行精準重複比對
        df_total['條碼_乾淨'] = df_total['條碼'].astype(str).str.replace('🔴 ', '').str.replace('🟢 ', '').str.strip()
        is_duplicate_barcode = df_total.duplicated(subset=['條碼_乾淨'], keep=False) & (~df_total['條碼_乾淨'].isin(["", "0", "nan", "None"]))
        
        df_anomaly = df_total[is_duplicate_barcode].sort_values(by=['條碼_乾淨', 'UID']).copy()
        df_anomaly = df_anomaly.drop(columns=['條碼_乾淨'], errors='ignore')
        df_total = df_total.drop(columns=['條碼_乾淨'], errors='ignore')
        
        if not df_anomaly.empty:
            def inject_emoji_alerts(row):
                remark = str(row.get('備註', ''))
                barcode = str(row.get('條碼', ''))
                if "名稱不同" in remark and not barcode.startswith("🔴"): row['條碼'] = f"🔴 {barcode}"
                elif "零售價不同" in remark and not barcode.startswith("🟢"): row['條碼'] = f"🟢 {barcode}"
                return row
            df_anomaly = df_anomaly.apply(inject_emoji_alerts, axis=1)
            
            st.warning("下方商品為系統抓出之條碼重複資料：🔴 代表名稱不一致，🟢 代表售價不一致。您可以直接在下方『備註』欄雙擊文字編寫 Note 紀錄！")
            
            edited_anomaly_df = st.data_editor(df_anomaly, use_container_width=True, disabled=[col for col in df_anomaly.columns if col not in ['move', '備註']], key="anomaly_editor")
            
            if st.button("🧹 執行審核與資料儲存", type="primary"):
                for index, edited_row in edited_anomaly_df.iterrows():
                    target_uid = edited_row['UID']
                    new_note = str(edited_row['備註']).strip().replace("🔴 ", "").replace("🟢 ", "")
                    df_total.loc[df_total['UID'] == target_uid, '備註'] = new_note

                uids_to_delete = edited_anomaly_df[edited_anomaly_df['move'] == True]['UID'].values
                if len(uids_to_delete) > 0:
                    df_to_delete = df_total[df_total['UID'].isin(uids_to_delete)].copy()
                    df_to_delete['刪除時間'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    df_to_delete = df_to_delete.drop(columns=['move'], errors='ignore')
                    
                    df_remaining = df_total[~df_total['UID'].isin(uids_to_delete)].copy()
                    df_remaining = df_remaining.drop(columns=['move'], errors='ignore')
                    df_remaining = run_cross_matching(df_remaining)
                    df_delete_log = pd.concat([df_delete_log, df_to_delete], ignore_index=True)
                    
                    if save_to_master_xlsm({"麗嬰國際產品總表": df_remaining, "刪除紀錄": df_delete_log}):
                        load_master_data.clear() 
                        st.session_state['merge_success_msg'] = "🧹 移轉封存與自訂 Note 備註已完整同步寫入雲端主資料庫！"
                        st.rerun()
                else:
                    df_remaining_only = run_cross_matching(df_total.drop(columns=['move'], errors='ignore'))
                    if save_to_master_xlsm({"麗嬰國際產品總表": df_remaining_only, "刪除紀錄": df_delete_log}):
                        load_master_data.clear() 
                        st.session_state['merge_success_msg'] = "📝 自訂 Note 備註內容已順利同步更新至雲端主資料庫！"
                        st.rerun()
        else:
            st.success("🟢 當前總表中沒有任何重複商品的衝突。")

    st.write("---")   


# -------------------------------------------------------------------------
# 子功能 5：📈 蝦皮商品清單轉換
# -------------------------------------------------------------------------
elif sub_page == "📈 蝦皮商品清單轉換":
    st.subheader("🛍️ 蝦皮賣場商品列表iSKU結構校正")
    
    # 🌟 需求 2：在頁面加入基本步驟指引
    with st.expander("📖 蝦皮資料校正與整合步驟指引（點擊展開）", expanded=True):
        st.markdown("""
        #### 💡 操作指引
        1. **下載蝦皮資料**：
           從 [蝦皮庫存列表](https://seller.shopee.tw/portal/product-mass/mass-update/download) 選擇模板 **「價格及庫存」** → 下載
           *(檔案命稱格式為: `mass_update_sales_info_xxxx_*.xlsx`)*
        2. **上傳該檔案並進行格式校正**：
           於下方上傳剛下載的 Excel 檔案。
        3. **以最新校正的蝦皮資料重新執行三表整合並回寫雲端**：
           校正完成後，點擊下方自動化推薦操作按鈕，即可一鍵完成回寫。
        """)
        
    st.write("---")
    
    # 延遲載入蝦皮快取
    df_shopee_history, df_shopee_current_list = load_shopee_data(ID_SHOPEE_MASTER)

    uploaded_shopee = st.file_uploader("📥 上傳新的蝦皮商品清單原始報表 (.xlsx/.xls/.xlsm) 進行格式校正：", type=["xlsx", "xls", "xlsm"], key="main_shopee_upload")
    
    if uploaded_shopee:
        file_bytes = uploaded_shopee.read()
        shopee_md5 = calculate_md5(file_bytes)
        
        # 🌟 需求 1 修復：取消「拒絕重複格式校正！系統已自動封鎖」的硬性阻斷
        is_duplicate = shopee_md5 in df_shopee_history['md5'].astype(str).values
        is_standardized = "shopee_standardized_" in uploaded_shopee.name

        if is_duplicate and not is_standardized:
            st.warning("⚠️ 系統偵測到此檔案先前似乎已處理過（MD5 重複）。為避免誤判，系統不再自動封鎖，您仍可點擊下方按鈕強制重新校正與整合。")
        elif is_standardized:
            st.info("ℹ️ 偵測到此為已校正過之標準檔案格式。")

        # 無論是否重複，都允許使用者點擊執行
        if st.button("🪄 執行蝦皮iSKU結構校正", type="primary", use_container_width=True):
            try:
                df_shopee_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, engine='openpyxl')
                if df_shopee_raw.shape[1] >= 11: df_shopee_raw.drop(df_shopee_raw.columns[10], axis=1, inplace=True)
                shopee_headers = df_shopee_raw.iloc[2].astype(str).str.strip().tolist()
                df_shopee = df_shopee_raw.iloc[6:].copy()
                df_shopee.columns = shopee_headers
                df_shopee.reset_index(drop=True, inplace=True)
                
                def calc_isku_row(row):
                    opt = str(row.get('商品選項貨號', '')).strip()
                    main = str(row.get('主商品貨號', '')).strip()
                    if opt in ["見選項", "null", "Null", "nan", "NaN", "None"]: opt = ""
                    if main in ["見選項", "null", "Null", "nan", "NaN", "None"]: main = ""
                    return opt if opt != "" else (main if main != "" else "蝦皮無iSKU")
                    
                df_shopee['iSKU'] = df_shopee.apply(calc_isku_row, axis=1)
                df_shopee['original_index'] = df_shopee.index
                cols_list = list(df_shopee.columns)
                if "iSKU" in cols_list and "價格" in cols_list:
                    cols_list.remove("iSKU")
                    cols_list.insert(cols_list.index("價格"), "iSKU")
                    df_shopee = df_shopee[cols_list]
                
                df_valid_isku = df_shopee[df_shopee['iSKU'] != "蝦皮無iSKU"].copy()
                df_isku_keep = df_valid_isku.sort_values(by=['iSKU', '價格', 'original_index']).drop_duplicates(subset=['iSKU'], keep='last')
                df_gtin_check = df_isku_keep.copy()
                df_gtin_check['GTIN_str'] = df_gtin_check['GTIN'].astype(str).str.strip().str.split('.').str[0]
                df_gtin_keep = df_gtin_check[~df_gtin_check['GTIN_str'].isin(["", "00", "0", "nan"])].sort_values(by=['GTIN_str', '價格', 'original_index']).drop_duplicates(subset=['GTIN_str'], keep='last')
                df_final_clean = pd.concat([df_gtin_keep, df_gtin_check[df_gtin_check['GTIN_str'].isin(["", "00", "0", "nan"])]]).sort_values(by='original_index')
                                 
                # 若非重複檔案，才寫入歷史紀錄避免 log 冗長
                if not is_duplicate:
                    new_hist_log = pd.DataFrame([{"檔案名稱": uploaded_shopee.name, "md5": shopee_md5, "匯入時間": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}])
                    df_shopee_history = pd.concat([df_shopee_history, new_hist_log], ignore_index=True)
                
                if save_to_shopee_master_xlsm({"蝦皮商品列表": df_final_clean, "匯入檔案": df_shopee_history}):
                    load_shopee_data.clear() 
                    get_cached_gdrive_id.clear()
                    st.session_state['shopee_clean'] = df_final_clean
                    st.success(f"🎉 蝦皮賣場商品列表iSKU結構校正完成！\n🟢 雲端現有主表 `蝦皮賣場商品列表.xlsm` 已成功同步覆寫更新！")
                    
            except Exception as e:
                st.error(f"讀取或清洗蝦皮檔案失敗: {str(e)}")

    # 確保自動化操作獨立於檔案上傳按鈕邏輯外，校正完成後隨時可見
    if 'shopee_clean' in st.session_state:
        st.markdown("---")
        st.markdown("### ⚡ 後續自動化推薦操作")
        
        # 🌟 對應需求 3：以最新校正的蝦皮資料重新執行三表整合
        if st.button("🚀 以最新校正的蝦皮資料重新執行三表整合並回寫雲端", type="primary", use_container_width=True):
            with st.spinner("⏳ 正在重新整理跨表聯結數據並回寫..."):
                if run_powerquery_and_update_gdrive():
                    st.success("✅ 成功！雲端『商品蝦皮麗嬰價格統整表』已同步使用最新校正後的蝦皮資料覆寫更新！")

        st.dataframe(st.session_state['shopee_clean'], use_container_width=True)
        towrite_shopee = io.BytesIO()
        st.session_state['shopee_clean'].to_excel(towrite_shopee, index=False)
        st.download_button(
            label="📥 下載此次iSKU校正蝦皮報表 (.xlsx)", 
            data=towrite_shopee.getvalue(), 
            file_name=f"蝦皮清洗完成對齊表_{datetime.date.today().strftime('%Y%m%d')}.xlsx", 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# -------------------------------------------------------------------------
# 子功能 6：🔀 sitegiant 採購入庫單轉換
# -------------------------------------------------------------------------
elif sub_page == "🔀 sitegiant 採購入庫單格式轉換":
    st.subheader("🛍️ SiteGiant 採購入庫單內容填寫")
    
    # ── 1. 基本設定與資料輸入 ──
    c_meta1, c_meta2 = st.columns(2)
    with c_meta1: 
        order_no = st.text_input("📝 請輸入訂單/銷貨單號：", value=datetime.date.today().strftime("%Y%m%d01"))
        vendor_options = ["麗嬰", "Buyee", "日亞", "其他"]
        selected_vendor = st.selectbox("🏬 請選擇採購廠商：", vendor_options, key="sg_vendor_selectbox")
        
        if selected_vendor == "其他":
            custom_vendor = st.text_input("✍️ 請輸入自訂廠商名稱：", key="sg_custom_vendor_name")
            vendor_name = custom_vendor if custom_vendor.strip() else "其他廠商"
        else:
            vendor_name = selected_vendor

    with c_meta2:
        recv_date = st.date_input("📅 選擇銷貨日期：", value=datetime.date.today())
        recv_date = recv_date.strftime("%y%m%d")
        
    st.write("---")

    # ── 2. 剪貼簿快速文字貼上區 (唯一入口，最穩定) ──
    st.markdown("### 1️⃣ 第一步：貼上原始資料")
    st.info("💡 請直接從 Excel 複製『國際條碼』與『數量』這兩欄資料，並貼入下方文字框中。")
    
    pasted_text = st.text_area(
        "📋 剪貼簿貼上區：", 
        height=150, 
        placeholder="請在此貼上...\n範例格式：\n4711234567890\t2\n4711234567891\t5"
    )
    
    if st.button("📥 解析並產生預覽表格", type="secondary", use_container_width=True):
        if pasted_text.strip():
            try:
                # 💡 自動解析 Tab 或空白分隔的 Excel 複製資料
                df_parsed = pd.read_csv(io.StringIO(pasted_text.strip()), sep=r'\s+|\t', engine='python', header=None, dtype=str)
                
                # 動態判斷使用者貼了幾欄
                col_count = df_parsed.shape[1]
                
                if col_count >= 2:
                    # 如果貼了兩欄以上，只取前兩欄
                    df_parsed = df_parsed.iloc[:, :2]
                    df_parsed.columns = ["國際條碼", "數量"]
                elif col_count == 1:
                    # 🌟 如果使用者『只貼了一欄（只有國際條碼）』，自動補上第二欄命名為數量，並預設帶入 1
                    df_parsed = df_parsed.iloc[:, :1]
                    df_parsed.columns = ["國際條碼"]
                    df_parsed["數量"] = "1"
                else:
                    raise ValueError("文字框內容為空或格式無法辨識。")
                
                # 將解析成功的資料存入 session_state 供下方表格渲染
                st.session_state['inward_input_df'] = df_parsed
                
                # 短暫顯示成功訊息後刷新畫面
                st.success(f"✅ 成功解析 {len(df_parsed)} 筆資料！已同步至下方預覽表格。")
                import time
                time.sleep(1.0)
                st.rerun()
            except Exception as parse_err:
                st.error(f"❌ 解析失敗，請確認複製內容格式是否正確：{parse_err}")
        else:
            st.warning("⚠️ 文字框為空，請先貼上資料。")

    st.write("---")

    # ── 3. 動態預覽與編輯區 ──
    st.markdown("### 2️⃣ 第二步：檢查與微調明細")
    
    # 確保 session_state 中有初始化的 DataFrame
    if 'inward_input_df' not in st.session_state:
        st.session_state['inward_input_df'] = pd.DataFrame(columns=["國際條碼", "數量"])
        
    # 渲染可編輯表格，並將修改結果同步回 session_state
    input_df = st.data_editor(
        st.session_state['inward_input_df'], 
        num_rows="dynamic", 
        use_container_width=True, 
        key="inward_grid"
    )
    st.session_state['inward_input_df'] = input_df 

    st.write("---") 
    
    # ── 4. 核心執行按鈕 ──
    st.markdown("### 3️⃣ 第三步：執行格式轉換")
    if st.button("✨ 執行貨品名稱和成本稅款導入並紀錄待處理商品", type="primary", use_container_width=True):
        if not order_no.strip() or input_df.empty: 
            st.error("❌ 轉換失敗！請填入銷貨單號，並確認上方表格有有效明細。")
        else:
            with st.spinner("正在由雲端獲取最新商品統整表並進行精準關聯..."):
                try:
                    if ID_PRICE_SUMMARY:
                        engine_kw = {"engine": "calamine"} if HAS_CALAMINE else {}
                        df_ref = pd.read_excel(download_gdrive_file_to_bytes(ID_PRICE_SUMMARY), **engine_kw)
                        df_ref['c_clean'] = df_ref['c'].astype(str).str.strip().str.split('.').str[0]
                            
                        result_rows = []
                        missing_items = []  # 收集所有異常商品
                        
                        for row in input_df.itertuples(index=False):
                            barcode_input = str(row.國際條碼).strip().split('.')[0] if pd.notna(row.國際條碼) else ""
                            if barcode_input in ["", "0", "nan", "None"]: continue
                            qty = int(row.數量) if pd.notna(row.數量) else 0
                        
                            # 預設狀態
                            sku_final = "⚠️ 提示：須新增iSKU"
                            prod_name = "⚠️ 未知商品名稱" 
                            category = ""
                            keywords = ""
                            cost_val = None 
                            tax_val = None
                            or_val = None
                            sal_val = None
                        
                            if not df_ref.empty and 'c_clean' in df_ref.columns:
                                match = df_ref[df_ref['c_clean'] == barcode_input]
                                if not match.empty:
                                    match_row = match.iloc[0]
                                    
                                    shopee_name = match_row.get('蝦皮商品名稱', None)
                                    internal_name = match_row.get('內部商品名稱', match_row.get('名稱', None)) 
                                    
                                    # 優先使用蝦皮名稱，若為空則使用內部名稱
                                    if pd.notna(shopee_name) and str(shopee_name).strip() not in ["", "nan", "None"]:
                                        prod_name = str(shopee_name).strip()
                                    elif pd.notna(internal_name) and str(internal_name).strip() not in ["", "nan", "None"]:
                                        prod_name = str(internal_name).strip()
                                    
                                    sku = match_row.get('自定義編碼', '')
                                    if pd.notna(sku) and str(sku).strip() != "":
                                        sku_final = str(sku).strip()
                                    
                                    category = match_row.get('分類定義', '')
                                    keywords = match_row.get('產品關鍵字', '')
                                
                                    # ── ⚙️ 廠商判定邏輯 ──
                                    if vendor_name == "麗嬰":
                                        def safe_float(v):
                                            try:
                                                return float(str(v).replace(',', '').strip())
                                            except (ValueError, TypeError):
                                                return None

                                        c_val = safe_float(match_row.get('麗嬰未稅價', None))
                                        t_val = safe_float(match_row.get('麗嬰稅款', None))
                                        s_val = safe_float(match_row.get('麗嬰批發含稅價', None))
                                        
                                        cost_val = round(c_val, 2) if c_val is not None else None
                                        tax_val = round(t_val, 2) if t_val is not None else None
                                        or_val = safe_float(match_row.get('麗嬰零售價', None))
                                        sal_val = round(s_val, 2) if s_val is not None else None

                            # 💡 判斷異常狀況並收集
                            issue_type = ""
                            if sku_final == "⚠️ 提示：須新增iSKU":
                                issue_type = "須建立自定義編碼"
                            elif prod_name == "⚠️ 未知商品名稱":
                                issue_type = "須建立賣場商品"
                                
                            if issue_type:
                                current_order_name = order_no if 'order_no' in locals() and str(order_no).strip() else "手動輸入未命名單號"
                                missing_items.append({
                                    "採購單檔名": current_order_name,
                                    "國際條碼": str(barcode_input).strip(),
                                    "狀況": issue_type,  
                                    "狀態": "待處理",    
                                    "建立時間": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                })

                            # 加入最終呈現的明細
                            result_rows.append({
                                "銷貨日期": str(recv_date), "國際條碼": barcode_input,
                                "庫存SKU": sku_final, "庫存貨品名稱": prod_name, 
                                "麗嬰零售價": or_val if vendor_name == "麗嬰" else None, 
                                "麗嬰批發含稅價": sal_val if vendor_name == "麗嬰" else None,
                                "成本": cost_val, "稅款": tax_val, "數量": qty,
                                "分類定義": category, "產品關鍵字": keywords
                            })
                            
                        # 💡 迴圈結束後，整批上傳至「尚未建立商品清單」
                        if missing_items:
                            try:
                                TARGET_SHEET_ID = "1Ixp9V_u2yU8hiWhxQCHNDB4kPKlxDGD2"
                                
                                try:
                                    raw_bytes = download_gdrive_file_to_bytes(TARGET_SHEET_ID)
                                    engine_kw = {"engine": "calamine"} if HAS_CALAMINE else {}
                                    df_missing = pd.read_excel(raw_bytes, sheet_name=0, dtype=str, **engine_kw)
                                except Exception as read_err:
                                    st.warning(f"⚠️ 無法讀取現有雲端清單內容，將重新建立。（錯誤訊息: {read_err}）")
                                    df_missing = pd.DataFrame(columns=["採購單檔名", "國際條碼", "狀況", "狀態", "建立時間"])
                                    
                                df_missing.columns = df_missing.columns.astype(str).str.strip()
                                
                                # 🌟 關鍵修改：同時組合「採購單檔名 + 國際條碼」作為唯一識別鍵 (Key)
                                if "採購單檔名" in df_missing.columns and "國際條碼" in df_missing.columns:
                                    existing_keys = set(
                                        df_missing["採購單檔名"].astype(str).str.strip() + "_" + 
                                        df_missing["國際條碼"].astype(str).str.strip()
                                    )
                                else:
                                    existing_keys = set()
                                
                                # 篩選出「同一張單號中尚未被記錄過的條碼」
                                new_missing_items = []
                                for item in missing_items:
                                    item_key = f"{item['採購單檔名']}_{item['國際條碼']}"
                                    if item_key not in existing_keys:
                                        new_missing_items.append(item)
                                        # 同時把剛加入的 key 補進 範圍內，避免這次送出的清單裡自己重複
                                        existing_keys.add(item_key)
                                
                                if new_missing_items:
                                    df_new_rows = pd.DataFrame(new_missing_items)
                                    df_missing = pd.concat([df_missing, df_new_rows], ignore_index=True)
                                    
                                    output_stream = io.BytesIO()
                                    with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
                                        df_missing.to_excel(writer, index=False, sheet_name="尚未建立商品清單")
                                    output_stream.seek(0)
                                    
                                    upload_or_update_gdrive_file(
                                        folder_id=None,
                                        file_name="尚未建立商品清單.xlsx", 
                                        file_bytes=output_stream.getvalue(),
                                        existing_file_id=TARGET_SHEET_ID
                                    )
                                    st.toast(f"🚨 已自動將 {len(new_missing_items)} 筆異常紀錄向下新增至雲端！", icon="⚠️")
                            except Exception as log_err:
                                st.error(f"⚠️ 自動記錄異常商品失敗: {str(log_err)}")
                                
                        if result_rows:
                            st.session_state['inward_result_df'] = pd.DataFrame(result_rows)
                            st.session_state['current_vendor_name'] = vendor_name
                            st.session_state['current_order_no'] = order_no
                            st.session_state['has_pending_items'] = len(missing_items) > 0 
                            st.success(f"🚀 格式勾稽完成！廠商已設定為：【{vendor_name}】")
                            st.rerun()
                            
                    else:
                        st.error("❌ 雲端找不到『商品蝦皮麗嬰價格統整表』。")
                        
                except Exception as e: 
                    st.error(f"❌ 錯誤: {str(e)}")

    # ── 5. 結果預覽與下載區 (執行後才會顯示) ──
    if 'inward_result_df' in st.session_state:
        st.write("---")
        res_df = st.session_state['inward_result_df']
        current_vendor = st.session_state.get('current_vendor_name', '未命名廠商')
        current_order = st.session_state.get('current_order_no', '0000')
        
        target_columns = ["國際條碼","庫存SKU", "庫存貨品名稱", "麗嬰零售價", "麗嬰批發含稅價", "成本", "稅款", "數量"]
        available_cols = [col for col in target_columns if col in res_df.columns]
        df_download = res_df[available_cols].copy()
        
        st.markdown(f"### 📋 【{current_vendor}】入庫明細結果預覽")
        st.info("💡 預覽表內的「麗嬰批發含稅價」、「成本」、「稅款」可雙擊進行最後微調，加總金額會即時變動。")
        
        # 渲染結果表
        edited_inward_df = st.data_editor(
            df_download,
            use_container_width=True,
            disabled=["國際條碼","庫存SKU", "庫存貨品名稱", "數量", "麗嬰零售價"],
            column_config={
                "麗嬰批發含稅價": st.column_config.NumberColumn(
                    "麗嬰批發含稅價", help="請手動輸入麗嬰批發含稅價", min_value=0.0, format="%.2f"
                ),
                "成本": st.column_config.NumberColumn(
                    "成本", help="請手動輸入未稅成本", min_value=0.0, format="%.2f"
                ),
                "稅款": st.column_config.NumberColumn(
                    "稅款", help="請手動輸入營業稅款", min_value=0.0, format="%.2f"
                ),
            },
            key="inward_items_editor_final"
        )
        
        # 成本與稅款動態加總
        h_cost = 0.0
        h_tax = 0.0
        for h_row in edited_inward_df.itertuples(index=False):
            h_qty = int(h_row.數量) if (hasattr(h_row, '數量') and pd.notna(h_row.數量)) else 0
            
            if hasattr(h_row, '成本') and pd.notna(h_row.成本):
                try: h_cost += float(h_row.成本) * h_qty
                except: pass
                
            if hasattr(h_row, '稅款') and pd.notna(h_row.稅款):
                try: h_tax += float(h_row.稅款) * h_qty
                except: pass
        
        st.markdown("#### 📊 本張單據入庫成本稅款即時統計看板")
        c_tot1, c_tot2 = st.columns(2)
        with c_tot1: 
            st.metric(label="💰 當前單據成本未稅總金額 (成本 * 數量)", value=f"$ {h_cost:,.2f} 元")
        with c_tot2: 
            st.metric(label="🧾 當前單據營業稅總金額 (稅款 * 數量)", value=f"$ {h_tax:,.2f} 元")
        
        st.write("---")
        
        # 下載與清除下一筆
        col_dl, col_reset = st.columns(2)
        with col_dl:
            towrite_inward = io.BytesIO()
            with pd.ExcelWriter(towrite_inward, engine='openpyxl') as writer:
                edited_inward_df.to_excel(writer, index=False, sheet_name="SiteGiant入庫單")
                
            has_pending = st.session_state.get('has_pending_items', False)
            pending_suffix = "_待處理" if has_pending else ""
            final_filename = f"sitegiant採購入庫單_{recv_date}_{current_vendor}_{current_order}{pending_suffix}.xlsx"

            st.download_button(
                label=f"📥 儲存並下載 SiteGiant 格式入庫單",
                data=towrite_inward.getvalue(),
                file_name=final_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
            
        with col_reset:
            if st.button("🔄 完成下載，清除畫面處理下一筆", type="secondary", use_container_width=True):
                # 清除所有的 state，恢復乾淨畫面
                if 'inward_result_df' in st.session_state:
                    del st.session_state['inward_result_df']
                st.session_state['inward_input_df'] = pd.DataFrame(columns=["國際條碼", "數量"])
                st.rerun()

# -------------------------------------------------------------------------
# 子功能 7：📜 sitegiant 歷史入庫單紀錄
# -------------------------------------------------------------------------
elif sub_page == "📜 sitegiant 歷史入庫單紀錄":
    st.subheader("📊 歷史入庫單與成本稅款加總指標檢視")
    hist_files = list_gdrive_files(ID_HISTORY_INWARD_FOLDER)
    if not hist_files: st.warning(f"💡 目前雲端無歷史單據。")
    else:
        file_options = {f['name']: f['id'] for f in hist_files}
        selected_hist_file = st.selectbox("🎯 選擇欲調閱的入庫對帳單：", list(file_options.keys()))
        if selected_hist_file:
            try:
                target_id = file_options[selected_hist_file]
                file_bytes = download_gdrive_file_to_bytes(target_id)
                df_hist_view = pd.read_excel(file_bytes, engine="calamine" if HAS_CALAMINE else None)
                st.markdown(f"📄 **當前雲端檔案**：`{selected_hist_file}` ｜ 📊 **單據品項數**：`{len(df_hist_view)} 筆`")
                st.dataframe(df_hist_view, use_container_width=True)
                
                h_cost = 0.0
                h_tax = 0.0
                for h_row in df_hist_view.itertuples(index=False):
                    h_qty = int(h_row.數量) if (hasattr(h_row, '數量') and pd.notna(h_row.數量)) else 0
                    
                    # 動態累加成本：先確認欄位存在且不是 NaN (空白)，才進行加總
                    if hasattr(h_row, '成本') and pd.notna(h_row.成本):
                        try: 
                            h_cost += float(h_row.成本) * h_qty
                        except ValueError: 
                            pass
                            
                    # 動態累加稅款：先確認欄位存在且不是 NaN (空白)，才進行加總
                    if hasattr(h_row, '稅款') and pd.notna(h_row.稅款):
                        try: 
                            h_tax += float(h_row.稅款) * h_qty
                        except ValueError: 
                            pass
                    
                st.markdown("#### 📊 本張單據入庫成本稅款")
                c_tot1, c_tot2 = st.columns(2)
                with c_tot1: 
                    st.metric(label="💰 成本未稅總金額 (成本 * 數量)", value=f"$ {h_cost:,.2f} 元")
                with c_tot2: 
                    st.metric(label="🧾 營業稅總金額 (稅款 * 數量)", value=f"$ {h_tax:,.2f} 元")    
                
                
                st.download_button(label="🔄 下載此歷史採購入庫單", data=file_bytes.getvalue(), file_name=selected_hist_file, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e: st.error(f"❌ 讀取失敗: {str(e)}")

# -------------------------------------------------------------------------
# 子功能 8：🔍 商品清單紀錄查詢 (支援自定義編碼與 c 欄位多筆查詢)
# -------------------------------------------------------------------------
elif sub_page == "🔍 商品清單紀錄查詢":
    st.subheader("📊 歷史商品清單紀錄查詢")
    hist_files = list_gdrive_files(ID_PROD_FOLDER)
    if not hist_files: 
        st.warning(f"💡 目前雲端無歷史單據。")
    else:
        file_options = {f['name']: f['id'] for f in hist_files}
        selected_hist_file = st.selectbox("🎯 選擇商品清單紀錄檔案：", list(file_options.keys()))
        
        if selected_hist_file:
            try:
                target_id = file_options[selected_hist_file]
                # 💡 導入快取與 calamine 高速讀取
                file_bytes = get_cached_gdrive_file_bytes(target_id)
                engine_kw = {"engine": "calamine"} if HAS_CALAMINE else {}
                df_hist_view = pd.read_excel(io.BytesIO(file_bytes), **engine_kw)
                
                st.markdown(f"📄 **當前雲端檔案**：`{selected_hist_file}` ｜ 📊 **iSKU總品項數**：`{len(df_hist_view)} 筆`")
                st.write("---")
                
                # ── 🌟 新增：多筆批次查詢區塊 ──
                st.markdown("#### 🔎 多筆批次編碼查詢")
                col_m, col_i = st.columns([1, 3])
                with col_m:
                    target_col = st.radio("選擇查詢依據欄位：", options=["自定義編碼", "c"], index=0)
                with col_i:
                    batch_input = st.text_area(f"請輸入多筆【{target_col}】（每筆請以換行、逗號或空格隔開）：", height=100, placeholder="例如：\n240101\n240102")
                
                df_display = df_hist_view.copy()
                if batch_input.strip() and target_col in df_display.columns:
                    import re
                    search_terms = [t.strip() for t in re.split(r'[\n,\s]+', batch_input) if t.strip()]
                    df_display[target_col] = df_display[target_col].astype(str).str.strip().str.split('.').str[0]
                    df_display = df_display[df_display[target_col].isin(search_terms)]
                    st.info(f"🎯 批次篩選結果：找到 **{len(df_display)}** 筆符合資料。")
                
                st.dataframe(df_display, use_container_width=True)
                
                towrite_prod = io.BytesIO()
                with pd.ExcelWriter(towrite_prod, engine='openpyxl') as writer:
                    df_display.to_excel(writer, index=False, sheet_name="商品清單查詢結果")
                
                st.download_button(
                    label="🔄 下載此歷史商品清單 (或篩選結果)", 
                    data=towrite_prod.getvalue(), 
                    file_name=selected_hist_file, 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e: 
                st.error(f"❌ 讀取失敗: {str(e)}")
# -------------------------------------------------------------------------
# 子功能 9：📦 SiteGiant 批量新增UPC
# -------------------------------------------------------------------------
elif sub_page == "📦 SiteGiant 批量新增UPC":
    st.subheader("📦 SiteGiant 批量新增UPC")
    
    # 顯示前置條件與後續操作
    with st.expander("📌 點擊查看操作流程與前/後置條件", expanded=True):
        st.markdown("""
        ### 📝 前置條件
        1. **確認 iSKU 對應 UPC**。
        2. **下載 Sitegiant 批量編輯 UPC 檔案**：
           - 至 Sitegiant 庫存列表 ➔ [批量編輯](https://sitegiant.co/items/batch-edit) UPC ➔ 下載 Excel。
           - 檔案命名格式通常為：`batch_edit_item_upc_assignment_all_DD-MM-YYYY-fileID`。
        3. **確認蝦皮賣場列表已校正**：
           - 需先執行過本系統的「蝦皮賣場商品列表iSKU結構校正」。
           - 本系統將自動調用雲端最新的校正表作為 UPC 比對資料庫。

        ### 🚀 後續操作
        - 將下方處理完畢並下載的新 Excel 檔案，上傳回 [Sitegiant 批量編輯](https://sitegiant.co/items/batch-edit) 覆蓋即可。
        """)

    # 1. 系統自動載入最新的蝦皮商品總表做為 VLOOKUP 的資料源
    df_shopee_hist, df_shopee_list = load_shopee_data(ID_SHOPEE_MASTER)

    if df_shopee_list.empty:
        st.error("❌ 無法從雲端讀取蝦皮商品列表！請先前往「蝦皮商品清單轉換」執行校正並回寫雲端。")
    else:
        st.info("✅ 系統已自動從雲端載入最新的蝦皮商品列表，準備好進行 UPC 交叉比對！")
        
        # 2. 讓使用者上傳從 Sitegiant 下載下來的原始檔 (新增 csv 支援)
        uploaded_sg_file = st.file_uploader("📥 請上傳 Sitegiant UPC 批量編輯下載檔 (.xlsx/.xls/.csv)", type=["xlsx", "xls", "csv"])
        
        if uploaded_sg_file:
            if st.button("⚡ 執行自動比對並填補 UPC", type="primary", use_container_width=True):
                with st.spinner("⏳ 正在自動比對蝦皮資料庫並填入缺失的 UPC..."):
                    try:
                        # 動態判斷讀取方式
                        if uploaded_sg_file.name.lower().endswith('.csv'):
                            df_sg = pd.read_csv(uploaded_sg_file, dtype=str)
                        else:
                            df_sg = pd.read_excel(uploaded_sg_file, dtype=str)
                        
                        # 💡 關鍵修正 1：強制清除所有標題欄位前後的隱藏空白與換行，避免比對失敗
                        df_sg.columns = df_sg.columns.astype(str).str.strip().str.replace('\n', '')
                        
                        # 智慧判斷 Sitegiant 的欄位名稱 (加入「庫存SKU」)
                        sg_sku_col = next((c for c in ["Item SKU", "商品 SKU", "SKU", "Item Sku", "item sku", "庫存SKU"] if c in df_sg.columns), None)
                        sg_upc_col = next((c for c in ["UPC", "國際條碼（UPC）", "國際條碼", "upc"] if c in df_sg.columns), None)
                        sg_main_col = next((c for c in ["Is Main UPC", "主要"] if c in df_sg.columns), None)
                        
                        if not sg_sku_col or not sg_upc_col:
                            # 💡 關鍵修正 2：如果找不到，直接把系統讀到的欄位名稱印出來抓漏！
                            st.error("❌ 檔案解析失敗：找不到對應的 SKU 或 UPC 欄位。")
                            st.warning(f"🕵️ 系統實際讀取到的所有欄位名稱為：\n`{list(df_sg.columns)}`")
                            st.info("💡 【除錯建議】：\n1. 請打開檔案確認，**標題列是否確實位於第一行**（若上方有空白列請先刪除）。\n2. 若您看到的欄位名稱與我們預設的不同，請告訴我，我們把它加入清單中！")
                        else:
                            # --- 下方原本的正常處理邏輯維持不變 ---
                            df_shopee_list['iSKU'] = df_shopee_list['iSKU'].astype(str).str.strip()
                            df_shopee_list['GTIN_str'] = df_shopee_list['GTIN'].astype(str).str.strip().str.split('.').str[0]
                            valid_shopee = df_shopee_list[~df_shopee_list['GTIN_str'].isin(["", "00", "0", "nan", "#N/A", "None", "空白"])]
                            
                            upc_map_exact = dict(zip(valid_shopee['iSKU'], valid_shopee['GTIN_str']))
                                                    
                            # 紀錄成功填補的索引 index
                            updated_indices = []
                            
                            # 逐筆比對與填補
                            for idx, row in df_sg.iterrows():
                                sku = str(row[sg_sku_col]).strip()
                                
                                # 💡 套用原本寫好的 clean_barcode 終極清洗，防止科學記號與空白
                                current_upc = clean_barcode(row.get(sg_upc_col, ""))
                                
                                # 判斷是否缺少 UPC (現在 current_upc 已經非常乾淨)
                                if current_upc in ["", "nan", "None", "0", "00"]:
                                    # 僅進行精準比對
                                    if sku in upc_map_exact:
                                        match_gtin = upc_map_exact[sku]
                                        
                                        # 填入 UPC
                                        df_sg.at[idx, sg_upc_col] = match_gtin
                                        
                                        # 自動將「主要」欄位設為 "是" 或 "Yes"
                                        if sg_main_col:
                                            df_sg.at[idx, sg_main_col] = "是" if "主要" in sg_main_col else "Yes"
                                        
                                        # 紀錄此筆為成功新增資料
                                        updated_indices.append(idx)
                            
                            # 💡 核心修改：僅保留成功填補 UPC 的資料列
                            df_sg_filtered = df_sg.loc[updated_indices].reset_index(drop=True)
                            
                            # 存入 Session State 供前端展示與下載
                            st.session_state['sg_upc_updated_df'] = df_sg_filtered
                            st.session_state['sg_upc_updated_count'] = len(df_sg_filtered)
                            
                            if len(df_sg_filtered) > 0:
                                st.success(f"🎉 處理完成！共成功自動填補 **{len(df_sg_filtered)}** 筆缺失的 UPC 資料。")
                            else:
                                st.warning("⚠️ 處理完成，但未找到任何可填補的缺失 UPC 資料。")
                            
                    except Exception as e:
                        st.error(f"❌ 處理檔案時發生錯誤：{str(e)}")
                        st.info("💡 請確認上傳的 SiteGiant 檔案格式與欄位內容。")

        # 3. 提供結果預覽與匯出下載 (僅含有填補成功者)
        if 'sg_upc_updated_df' in st.session_state and not st.session_state['sg_upc_updated_df'].empty:
            df_result = st.session_state['sg_upc_updated_df']
            
            st.markdown(f"### 📋 成功新增 UPC 預覽（共 {len(df_result)} 筆）")
            st.dataframe(df_result, use_container_width=True)
            
            towrite_sg = io.BytesIO()
            with pd.ExcelWriter(towrite_sg, engine='openpyxl') as writer:
                df_result.to_excel(writer, index=False)
                
            download_filename = f"batch_edit_upc_added_only_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
            
            st.download_button(
                label=f"📥 下載包含 {len(df_result)} 筆成功自動填補的 SiteGiant 檔案",
                data=towrite_sg.getvalue(),
                file_name=download_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
# -------------------------------------------------------------------------
# 子功能 10：📋 採購單待處理
# -------------------------------------------------------------------------
elif sub_page == "📋 採購單待處理":
    st.subheader("📋 採購單待處理 (尚未建立商品清單)")
    
    # 目標雲端檔案 ID
    TARGET_SHEET_ID = "1Ixp9V_u2yU8hiWhxQCHNDB4kPKlxDGD2"
    
    with st.spinner("⏳ 正在由雲端獲取待處理清單..."):
        try:
            # 💡 導入快取與 calamine 高速讀取
            file_bytes = get_cached_gdrive_file_bytes(TARGET_SHEET_ID)
            engine_kw = {"engine": "calamine"} if HAS_CALAMINE else {}
            df_pending = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, dtype=str, **engine_kw)
            
            # 清理可能有空白的標題
            df_pending.columns = df_pending.columns.astype(str).str.strip()
            
            if df_pending.empty:
                st.success("🎉 太棒了！目前沒有任何待處理的異常商品與採購單。")
            else:
                st.markdown(f"📊 **目前待處理總筆數**：`{len(df_pending)} 筆`")
                st.info("💡 下方為過去轉換入庫單時，找不到雲端統整表對應紀錄的異常商品，請調閱並盡速前往建立或更新資料。")
                
                # 在畫面上渲染預覽表
                st.dataframe(df_pending, use_container_width=True)
                
                # 提供下載最新清單的功能
                towrite_pending = io.BytesIO()
                with pd.ExcelWriter(towrite_pending, engine='openpyxl') as writer:
                    df_pending.to_excel(writer, index=False, sheet_name="尚未建立商品清單")
                
                st.download_button(
                    label="📥 下載待處理清單報表 (.xlsx)",
                    data=towrite_pending.getvalue(),
                    file_name=f"尚未建立商品清單_匯出_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                # (選用) 加入重新載入按鈕，方便使用者若手動去雲端改完，可以馬上刷新快取
                if st.button("🔄 重新載入最新雲端資料", use_container_width=True):
                    get_cached_gdrive_file_bytes.clear()
                    st.rerun()
                    
        except Exception as e:
            st.error("❌ 讀取雲端清單失敗，可能是該檔案尚未被系統自動建立或權限不足。")
            st.warning(f"🔍 系統詳細報錯訊息：`{str(e)}`")  # 👈 加入這行可看見實際 API 錯誤代碼 (如 404/403)
